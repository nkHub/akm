"""Agent 工作区文件工具与安全边界测试。"""

import asyncio
import json
from pathlib import Path

import pytest

from akm.agent_runtime.tools import (
    build_workspace_tools,
    _safe_resolve_workspace_path,
    _workspace_root,
    reset_request_workspace_root,
    set_request_workspace_root,
)

# 供测试复用的工作区配置
WORKSPACE = "/tmp/akm-test-workspace"


def _workspace_cfg(**overrides):
    """构造工作区相关配置 dict，覆盖默认值。"""
    cfg = {
        "agent_workspace_root": WORKSPACE,
        "agent_write_tools_enabled": True,
        "agent_run_shell_enabled": True,
    }
    cfg.update(overrides)
    return cfg


def _handlers(monkeypatch, **cfg):
    """注册工作区工具并返回 {name: handler}，配置从 cfg 生成。"""
    monkeypatch.setattr(
        "akm.agent_runtime.tools.load_config", lambda: _workspace_cfg(**cfg)
    )
    return {t.name: t.handler for t in build_workspace_tools()}


@pytest.fixture(autouse=True)
def _workspace(tmp_path, monkeypatch):
    """每个测试使用独立临时目录作为工作区，避免跨测试污染。"""
    root = tmp_path / "ws"
    root.mkdir()
    monkeypatch.setattr("akm.agent_runtime.tools._workspace_root", lambda: root.resolve())
    yield root


# ── 路径沙箱 ──


def test_safe_resolve_rejects_outside_absolute_path():
    """绝对路径越界应被拒绝。"""
    with pytest.raises(ValueError, match="超出工作区"):
        _safe_resolve_workspace_path("/etc/passwd")


def test_safe_resolve_rejects_traversal():
    """相对路径中的 .. 穿越应被拒绝。"""
    with pytest.raises(ValueError, match="超出工作区"):
        _safe_resolve_workspace_path("../../etc/passwd")


def test_safe_resolve_accepts_inside_path():
    """工作区内的相对/绝对路径应被正常解析。"""
    p = _safe_resolve_workspace_path("sub/file.txt", must_exist=False)
    assert str(p).endswith("/sub/file.txt")


def test_safe_resolve_rejects_symlink_escape(tmp_path):
    """软链接指向工作区外时应被拒绝（resolve 后越界）。"""
    outside = tmp_path / "outside.txt"
    outside.write_text("secret")
    link = tmp_path / "ws" / "link.txt"
    link.symlink_to(outside)
    with pytest.raises(ValueError, match="超出工作区"):
        _safe_resolve_workspace_path("link.txt")


def test_workspace_root_returns_none_when_unconfigured(monkeypatch):
    """未配置 agent_workspace_root 时 _workspace_root 应返回 None。"""
    monkeypatch.setattr(
        "akm.agent_runtime.tools.load_config", lambda: {"agent_workspace_root": ""}
    )
    assert _workspace_root() is None


def test_request_workspace_root_override(monkeypatch):
    """请求级 workspace_root 只能选择全局工作区内的子目录。"""
    monkeypatch.setattr(
        "akm.agent_runtime.tools.load_config", lambda: {"agent_workspace_root": "/global/ws"}
    )
    assert str(_workspace_root()) == str(Path("/global/ws").resolve())

    token = set_request_workspace_root("/global/ws/project")
    try:
        assert str(_workspace_root()) == str(Path("/global/ws/project").resolve())
    finally:
        reset_request_workspace_root(token)

    assert str(_workspace_root()) == str(Path("/global/ws").resolve())


def test_request_workspace_root_rejects_outside_global_root(monkeypatch):
    """请求级工作区不得把文件工具的沙箱范围扩大到全局根目录之外。"""
    monkeypatch.setattr(
        "akm.agent_runtime.tools.load_config", lambda: {"agent_workspace_root": "/global/ws"}
    )
    token = set_request_workspace_root("/other/ws")
    try:
        with pytest.raises(ValueError, match="必须位于"):
            _workspace_root()
    finally:
        reset_request_workspace_root(token)


def test_request_workspace_root_empty_keeps_global(monkeypatch):
    """空字符串的请求级覆盖不应生效，仍使用全局配置。"""
    monkeypatch.setattr(
        "akm.agent_runtime.tools.load_config", lambda: {"agent_workspace_root": "/global/ws"}
    )
    token = set_request_workspace_root("")
    try:
        assert str(_workspace_root()) == str(Path("/global/ws").resolve())
    finally:
        reset_request_workspace_root(token)


# ── 读工具 ──


@pytest.mark.asyncio
async def test_read_file_within_workspace(_workspace, monkeypatch):
    """读取工作区内文件应返回内容与行信息。"""
    f = _workspace / "hello.txt"
    f.write_text("第一行\n第二行\n第三行\n", encoding="utf-8")
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_read_file"](path="hello.txt"))

    assert out["path"].endswith("/hello.txt")
    assert "第一行" in out["content"]
    assert out["start_line"] == 0


@pytest.mark.asyncio
async def test_read_file_outside_workspace_rejected(_workspace, monkeypatch):
    """读取工作区外文件应返回错误，而不是泄露内容。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_read_file"](path="/etc/passwd"))

    assert "error" in out
    assert "超出工作区" in out["error"]


@pytest.mark.asyncio
async def test_read_file_missing_returns_error(_workspace, monkeypatch):
    """读取不存在的文件应返回错误。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_read_file"](path="nope.txt"))

    assert "error" in out


@pytest.mark.asyncio
async def test_list_dir_shows_entries(_workspace, monkeypatch):
    """list_dir 应返回工作区内的目录条目。"""
    (_workspace / "a.txt").write_text("x", encoding="utf-8")
    (_workspace / "sub").mkdir()
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_list_dir"](path=""))

    names = {e["name"] for e in out["entries"]}
    assert "a.txt" in names
    assert "sub" in names


@pytest.mark.asyncio
async def test_glob_matches_within_workspace(_workspace, monkeypatch):
    """glob 应返回相对路径且不允许越界模式。"""
    (_workspace / "main.py").write_text("", encoding="utf-8")
    (_workspace / "src").mkdir()
    (_workspace / "src" / "lib.py").write_text("", encoding="utf-8")
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_glob"](pattern="**/*.py"))
    assert "main.py" in out["matches"]
    assert "src/lib.py" in out["matches"]

    bad = json.loads(await handlers["akm_glob"](pattern="../**"))
    assert "error" in bad


@pytest.mark.asyncio
async def test_grep_searches_content(_workspace, monkeypatch):
    """grep 应返回命中文件、行号与行内容。"""
    (_workspace / "code.py").write_text("def foo():\n    return 1\n", encoding="utf-8")
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_grep"](pattern="def foo"))

    assert out["total"] == 1
    assert out["results"][0]["file"] == "code.py"
    assert out["results"][0]["line"] == 1


@pytest.mark.asyncio
async def test_grep_skips_symlink_outside_workspace(_workspace, monkeypatch, tmp_path):
    """grep 必须逐项校验软链接真实路径，不能读取工作区外的文件内容。"""
    outside = tmp_path / "outside.txt"
    outside.write_text("secret marker", encoding="utf-8")
    (_workspace / "outside-link.txt").symlink_to(outside)
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_grep"](pattern="secret marker"))

    assert out["total"] == 0


@pytest.mark.asyncio
async def test_file_info_returns_metadata(_workspace, monkeypatch):
    """file_info 应返回类型、大小与修改时间。"""
    f = _workspace / "meta.txt"
    f.write_text("hello", encoding="utf-8")
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_file_info"](path="meta.txt"))

    assert out["type"] == "file"
    assert out["size"] == 5


# ── 写工具开关 ──


@pytest.mark.asyncio
async def test_write_tool_requires_enabled_flag(_workspace, monkeypatch):
    """agent_write_tools_enabled=false 时写工具不注册（模型不可见即不可调）。"""
    handlers = _handlers(monkeypatch, agent_write_tools_enabled=False)

    assert "akm_write_file" not in handlers
    assert "akm_edit_file" not in handlers


@pytest.mark.asyncio
async def test_write_file_and_edit(_workspace, monkeypatch):
    """写文件并编辑：覆盖写入、精确替换。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_write_file"](path="code.py", content="print('old')\n"))
    assert out["ok"] is True
    assert (_workspace / "code.py").read_text(encoding="utf-8") == "print('old')\n"

    out = json.loads(await handlers["akm_edit_file"](path="code.py", old_string="old", new_string="new"))
    assert out["replaced"] == 1
    assert (_workspace / "code.py").read_text(encoding="utf-8") == "print('new')\n"


# ── 结构化编辑（行号模式）──


@pytest.mark.asyncio
async def test_edit_file_line_mode_replaces_range(_workspace, monkeypatch):
    """行号模式应按行区间整体替换并返回行数变化。"""
    f = _workspace / "code.py"
    f.write_text("line1\nline2\nline3\nline4\n", encoding="utf-8")
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_edit_file"](path="code.py", start_line=2, end_line=3, new_content="A\nB"))

    assert out["ok"] is True
    assert out["old_lines"] == 2
    assert out["new_lines"] == 2
    assert f.read_text(encoding="utf-8") == "line1\nA\nB\nline4\n"


@pytest.mark.asyncio
async def test_edit_file_line_mode_single_line_without_end(_workspace, monkeypatch):
    """行号模式不传 end_line 时应只替换一行，且保留尾部换行风格。"""
    f = _workspace / "code.py"
    f.write_text("line1\nline2\nline3\n", encoding="utf-8")
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_edit_file"](path="code.py", start_line=2, new_content="two"))

    assert out["ok"] is True
    assert out["old_lines"] == 1
    assert f.read_text(encoding="utf-8") == "line1\ntwo\nline3\n"


@pytest.mark.asyncio
async def test_edit_file_line_mode_anchor_mismatch_rejected(_workspace, monkeypatch):
    """行号模式锚点校验失败应拒绝编辑，防止行号漂移后改错位置。"""
    f = _workspace / "code.py"
    f.write_text("line1\nline2\nline3\n", encoding="utf-8")
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_edit_file"](path="code.py", start_line=2, end_line=3, old_string="不存在的锚点", new_content="x"))

    assert "error" in out
    assert "锚点校验失败" in out["error"]


@pytest.mark.asyncio
async def test_edit_file_line_mode_out_of_range_rejected(_workspace, monkeypatch):
    """行号越界应被拒绝并给出文件实际行数。"""
    f = _workspace / "code.py"
    f.write_text("a\nb\n", encoding="utf-8")
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_edit_file"](path="code.py", start_line=5, new_content="x"))

    assert "error" in out
    assert "越界" in out["error"]


@pytest.mark.asyncio
async def test_edit_file_content_mode_still_works(_workspace, monkeypatch):
    """内容模式（旧参数）应保持原有替换行为。"""
    f = _workspace / "code.py"
    f.write_text("print('old')\n", encoding="utf-8")
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_edit_file"](path="code.py", old_string="old", new_string="new"))

    assert out["replaced"] == 1
    assert f.read_text(encoding="utf-8") == "print('new')\n"


@pytest.mark.asyncio
async def test_write_file_rejects_outside_workspace(_workspace, monkeypatch):
    """写文件到工作区外应被拒绝。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_write_file"](path="/tmp/evil.txt", content="x"))

    assert "error" in out


@pytest.mark.asyncio
async def test_delete_rejects_workspace_root(_workspace, monkeypatch):
    """删除工作区根目录应被拒绝。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_delete_file"](path=".", recursive=True))

    assert "error" in out
    assert "工作区根目录" in out["error"]


@pytest.mark.asyncio
async def test_delete_rejects_directory_without_recursive(_workspace, monkeypatch):
    """删除目录但未设置 recursive=true 应被拒绝（目录完好）。"""
    handlers = _handlers(monkeypatch)

    d = _workspace / "adir"
    d.mkdir()
    (d / "f1.txt").write_text("x", encoding="utf-8")

    out = json.loads(await handlers["akm_delete_file"](path="adir"))

    assert "error" in out
    assert "recursive=true" in out["error"]
    assert d.is_dir()  # 目录及其内容完好


@pytest.mark.asyncio
async def test_delete_directory_recursive(_workspace, monkeypatch):
    """recursive=true 时递归删除目录及其所有内容。"""
    handlers = _handlers(monkeypatch)

    d = _workspace / "adir"
    d.mkdir()
    (d / "sub").mkdir()
    (d / "f1.txt").write_text("x", encoding="utf-8")
    (d / "sub" / "f2.txt").write_text("y", encoding="utf-8")

    out = json.loads(await handlers["akm_delete_file"](path="adir", recursive=True))

    assert out["ok"] is True
    assert not d.exists()  # 目录连同子目录与文件全部删除


@pytest.mark.asyncio
async def test_make_dir_and_delete(_workspace, monkeypatch):
    """创建目录并删除文件。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_make_dir"](path="a/b/c"))
    assert out["ok"] is True
    assert (_workspace / "a" / "b" / "c").is_dir()

    f = _workspace / "a" / "tmp.txt"
    f.write_text("x", encoding="utf-8")
    out = json.loads(await handlers["akm_delete_file"](path="a/tmp.txt"))
    assert out["ok"] is True
    assert not f.exists()


# ── shell 工具开关 ──


@pytest.mark.asyncio
async def test_run_shell_requires_enabled_flag(_workspace, monkeypatch):
    """agent_run_shell_enabled=false 时 shell 工具不注册。"""
    handlers = _handlers(monkeypatch, agent_run_shell_enabled=False)

    assert "akm_run_shell" not in handlers


@pytest.mark.asyncio
async def test_run_shell_executes_with_workspace_cwd(_workspace, monkeypatch):
    """shell 工具直接执行命令字符串，且以工作区为 cwd。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_run_shell"](command="pwd && echo ok"))

    assert out["exit_code"] == 0
    assert str(_workspace) in out["output"]
    assert "ok" in out["output"]


@pytest.mark.asyncio
async def test_run_shell_timeout(_workspace, monkeypatch):
    """shell 命令超时应被终止并返回错误。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_run_shell"](command="sleep 5", timeout=1))

    assert "超时" in out["error"]


# ── git 工具 ──


@pytest.mark.asyncio
async def test_run_git_requires_enabled_flag(_workspace, monkeypatch):
    """agent_git_enabled=false 时 git 工具不注册。"""
    handlers = _handlers(monkeypatch, agent_git_enabled=False)

    assert "akm_run_git" not in handlers


@pytest.mark.asyncio
async def test_run_git_status_in_workspace(_workspace, monkeypatch):
    """git 工具应以工作区为 cwd 执行 git 命令并返回输出与退出码。"""
    import subprocess as _sp

    _sp.run(["git", "init"], cwd=str(_workspace), capture_output=True, text=True)
    (_workspace / "a.txt").write_text("x", encoding="utf-8")
    handlers = _handlers(monkeypatch, agent_git_enabled=True)

    out = json.loads(await handlers["akm_run_git"](operation="status"))

    assert out["exit_code"] == 0
    assert "a.txt" in out["output"]


@pytest.mark.asyncio
async def test_run_git_rejects_unknown_operation(_workspace, monkeypatch):
    """Git 工具只支持服务端定义的结构化操作。"""
    handlers = _handlers(monkeypatch, agent_git_enabled=True)

    out = json.loads(await handlers["akm_run_git"](operation="clone"))

    assert "error" in out
    assert "不支持" in out["error"]


@pytest.mark.asyncio
async def test_run_git_rejects_paths_outside_workspace(_workspace, monkeypatch):
    """Git 路径参数不能通过绝对路径或 .. 逃离工作区。"""
    handlers = _handlers(monkeypatch, agent_git_enabled=True)

    out = json.loads(await handlers["akm_run_git"](operation="add", paths=["../outside.txt"]))

    assert "error" in out
    assert "工作区内的相对路径" in out["error"]


@pytest.mark.asyncio
async def test_run_git_commit_requires_message(_workspace, monkeypatch):
    """提交操作必须提供受限长度的提交说明。"""
    handlers = _handlers(monkeypatch, agent_git_enabled=True)

    out = json.loads(await handlers["akm_run_git"](operation="commit"))

    assert "error" in out
    assert "message" in out["error"]


# ── 工具注册集合 ──


def test_workspace_tools_registration(monkeypatch):
    """写工具、shell 工具与 git 工具按开关注册；读工具始终注册。"""
    monkeypatch.setattr(
        "akm.agent_runtime.tools.load_config",
        lambda: {
            "agent_workspace_root": WORKSPACE,
            "agent_write_tools_enabled": False,
            "agent_run_shell_enabled": False,
            "agent_git_enabled": False,
        },
    )
    names = {t.name for t in build_workspace_tools()}
    assert {"akm_read_file", "akm_list_dir", "akm_glob", "akm_grep", "akm_file_info"} <= names
    assert "akm_write_file" not in names
    assert "akm_xlsx" not in names
    assert "akm_run_shell" not in names
    assert "akm_run_git" not in names

    monkeypatch.setattr(
        "akm.agent_runtime.tools.load_config",
        lambda: {
            "agent_workspace_root": WORKSPACE,
            "agent_write_tools_enabled": True,
            "agent_run_shell_enabled": True,
            "agent_git_enabled": True,
        },
    )
    names = {t.name for t in build_workspace_tools()}
    assert "akm_write_file" in names
    assert "akm_edit_file" in names
    assert "akm_make_dir" in names
    assert "akm_delete_file" in names
    assert "akm_xlsx" in names
    assert "akm_run_shell" in names
    assert "akm_run_git" in names


def test_shell_and_git_schema(monkeypatch):
    """shell 工具暴露 command 字符串；git 工具保持固定 operation，不暴露自由命令。"""
    monkeypatch.setattr(
        "akm.agent_runtime.tools.load_config",
        lambda: _workspace_cfg(agent_git_enabled=True),
    )
    tools = {tool.name: tool for tool in build_workspace_tools()}

    shell_properties = tools["akm_run_shell"].parameters["properties"]
    git_properties = tools["akm_run_git"].parameters["properties"]
    assert "command" in shell_properties
    assert shell_properties["command"]["type"] == "string"
    assert "operation" in git_properties
    assert "command" not in git_properties


# ── xlsx 电子表格工具 ──


@pytest.mark.asyncio
async def test_xlsx_create_and_verify_content(_workspace, monkeypatch):
    """akm_xlsx create 生成可被 openpyxl 读回的 xlsx，数据逐行落位。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_xlsx"](
        action="create",
        path="book.xlsx",
        data=[["name", "age"], ["alice", 30], ["bob", 25]],
    ))

    assert out.get("ok") is True
    assert out["action"] == "create"
    assert (Path(str(out["path"]))).is_file()
    from openpyxl import load_workbook
    wb = load_workbook(str(out["path"]))
    ws = wb["Sheet1"]
    assert [c.value for c in ws[1]] == ["name", "age"]
    assert ws["A2"].value == "alice"
    assert ws["B2"].value == 30


@pytest.mark.asyncio
async def test_xlsx_create_multiple_sheets(_workspace, monkeypatch):
    """create 的 data 传 {sheet名: 二维数组} 时生成多个工作表。"""
    handlers = _handlers(monkeypatch)

    out = json.loads(await handlers["akm_xlsx"](
        action="create",
        path="multi.xlsx",
        data={"Sheet1": [["a"]], "Sheet2": [["b", 2]]},
    ))

    assert out.get("ok") is True
    from openpyxl import load_workbook
    wb = load_workbook(str(out["path"]))
    assert wb.sheetnames == ["Sheet1", "Sheet2"]
    assert wb["Sheet2"]["A1"].value == "b"


@pytest.mark.asyncio
async def test_xlsx_edit_writes_cells(_workspace, monkeypatch):
    """edit 按 updates 写单元格；缺 sheet 默认 Sheet1。"""
    handlers = _handlers(monkeypatch)

    await handlers["akm_xlsx"](action="create", path="book.xlsx", data=[["x"]])
    out = json.loads(await handlers["akm_xlsx"](
        action="edit",
        path="book.xlsx",
        updates=[
            {"sheet": "Sheet1", "cell": "B1", "value": "score"},
            {"cell": "B2", "value": 99},
        ],
    ))

    assert out.get("ok") is True
    assert out["updated"] == 2
    from openpyxl import load_workbook
    wb = load_workbook(str(out["path"]))
    assert wb["Sheet1"]["B1"].value == "score"
    assert wb["Sheet1"]["B2"].value == 99


@pytest.mark.asyncio
async def test_xlsx_create_rejects_existing_without_overwrite(_workspace, monkeypatch):
    """目标已存在且未传 overwrite=true 时应拒绝覆盖。"""
    handlers = _handlers(monkeypatch)
    await handlers["akm_xlsx"](action="create", path="book.xlsx", data=[["x"]])

    out = json.loads(await handlers["akm_xlsx"](action="create", path="book.xlsx", data=[["y"]]))

    assert "error" in out
    assert "已存在" in out["error"]


@pytest.mark.asyncio
async def test_xlsx_rejects_path_outside_workspace(_workspace, monkeypatch):
    """xlsx 路径越界（绝对路径/..穿越）应被拒绝。"""
    handlers = _handlers(monkeypatch)

    for path in ("/etc/passwd", "../outside.xlsx"):
        out = json.loads(await handlers["akm_xlsx"](action="create", path=path, data=[["x"]]))
        assert "error" in out
        assert "超出工作区" in out["error"]


@pytest.mark.asyncio
async def test_xlsx_rejects_bad_action_and_missing_args(_workspace, monkeypatch):
    """非法 action、create 缺 data、edit 缺 updates 应返回错误。"""
    handlers = _handlers(monkeypatch)

    bad_action = json.loads(await handlers["akm_xlsx"](action="delete", path="a.xlsx"))
    assert "error" in bad_action
    assert "action" in bad_action["error"]

    no_data = json.loads(await handlers["akm_xlsx"](action="create", path="a.xlsx"))
    assert "error" in no_data

    no_updates = json.loads(await handlers["akm_xlsx"](action="edit", path="a.xlsx"))
    assert "error" in no_updates


def test_xlsx_schema(monkeypatch):
    """xlsx 工具注册在写工具开关下，暴露 action/path/data/updates 参数。"""
    monkeypatch.setattr(
        "akm.agent_runtime.tools.load_config",
        lambda: _workspace_cfg(),
    )
    tools = {tool.name: tool for tool in build_workspace_tools()}

    assert "akm_xlsx" in tools
    props = tools["akm_xlsx"].parameters["properties"]
    assert props["action"]["enum"] == ["create", "edit"]
    assert "path" in props
    assert "data" in props
    assert "updates" in props
