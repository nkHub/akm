"""供 Agent Loop 使用的 AKM 内置只读调试工具。"""

import asyncio
import base64
import ctypes
import datetime
import json
import logging
import mimetypes
import os
import re
import shutil
import smtplib
import subprocess
import tempfile
import uuid
from contextvars import ContextVar
from email.message import EmailMessage
from email.utils import formataddr, make_msgid, parseaddr
from pathlib import Path
from typing import Any

import httpx
from fastapi import FastAPI

from akm.agent_runtime.loop import ToolDef
from akm.agent_runtime.tavily_mcp import tavily_search
from akm.audit import list_logs_async
from akm.config import load_config
from akm.key_pool import key_model_list, list_keys

logger = logging.getLogger(__name__)

# 文件读取工具单次返回的最大字节数，防止超大文件撑爆模型上下文
_WORKSPACE_READ_MAX_BYTES = 60000
# grep 工具单次返回的最大命中行数
_WORKSPACE_GREP_MAX_RESULTS = 100
# 目录工具单次返回的最大条目数
_WORKSPACE_DIR_MAX_RESULTS = 500
# 图片生成/编辑单次允许请求的最大张数，限制异常参数造成的上游费用放大
_AGENT_IMAGE_MAX_COUNT = 4
# 图片编辑单个本地文件或 base64 解码后的最大字节数，避免工具调用占满进程内存。
_AGENT_IMAGE_MAX_INPUT_BYTES = 20 * 1024 * 1024
# run_shell 工具默认超时（秒）
_WORKSPACE_SHELL_TIMEOUT_SEC = 60
# 读文件工具允许读取的最大文件大小（st_size 预检），超过时拒绝读取，
# 防止把超大文件整个读入内存拖慢进程。文本源码/配置文件通常远小于此值。
_WORKSPACE_READ_MAX_FILE_BYTES = 50 * 1024 * 1024
# grep 单文件扫描上限：超过该字节数的文件直接跳过（不逐行正则匹配）
_WORKSPACE_GREP_MAX_FILE_BYTES = 10 * 1024 * 1024
# grep 单行最大长度：超过则跳过该行，避免把整文件内容聚合成单行后
# 正则回退爆炸（ReDoS）或匹配超大行
_WORKSPACE_GREP_MAX_LINE_BYTES = 1024 * 1024
# grep 单次扫描的全局字节预算：达到后停止继续遍历，防止全盘递归拖垮进程
_WORKSPACE_GREP_MAX_SCAN_BYTES = 128 * 1024 * 1024
# 写文件工具允许写入的内容大小上限，防止模型一次写入 GB 级内容撑爆磁盘/内存
_WORKSPACE_WRITE_MAX_BYTES = 10 * 1024 * 1024

# 请求级工作区覆盖：由 AgentLoop 在单次请求执行工具期间设置，优先于
# config.json 的全局 agent_workspace_root；空字符串表示不使用覆盖。
# 用 ContextVar 保证并发请求之间互不影响（每个请求独立的 asyncio 上下文）。
_request_workspace_root: ContextVar[str | None] = ContextVar("agent_request_workspace_root", default=None)


def set_request_workspace_root(root: str):
    """设置本次请求执行工具期间的工作区覆盖，返回用于恢复的 token。

    Args:
        root: 覆盖的工作区根目录绝对路径；空字符串表示不覆盖（走全局配置）。
    """
    return _request_workspace_root.set(str(root or "").strip())


def reset_request_workspace_root(token) -> None:
    """恢复 set_request_workspace_root 之前的工作区上下文。"""
    _request_workspace_root.reset(token)


def _workspace_root() -> Path | None:
    """返回当前请求的 Agent 工作区沙箱根目录（已展开 ~、已 resolve），未配置时返回 None。

    工作区根目录优先取请求级覆盖（/v1/agent 请求的 workspace_root 字段，
    由 AgentLoop 执行工具前设置），否则回退 config.json 的全局
    agent_workspace_root。所有文件读写工具都只能访问该目录内的路径
    （防止路径穿越读写任意文件）；未配置时文件工具整体不可用。
    """
    configured_raw = str(load_config().get("agent_workspace_root") or "").strip()
    configured_root = Path(configured_raw).expanduser().resolve() if configured_raw else None
    request_raw = _request_workspace_root.get()
    if request_raw is None or not request_raw.strip():
        return configured_root
    if configured_root is None:
        raise ValueError("未配置 agent_workspace_root，请求级工作区不可用")
    requested_root = Path(request_raw).expanduser().resolve()
    try:
        requested_root.relative_to(configured_root)
    except ValueError:
        raise ValueError("请求级 workspace_root 必须位于 agent_workspace_root 内")
    return requested_root


def _path_is_under(path: Path, root: Path) -> bool:
    """判断已 resolve 的路径是否位于指定根目录内。"""
    try:
        path.relative_to(root)
    except ValueError:
        return False
    return True


def _allowed_image_path(path: Path) -> bool:
    """校验 Agent 请求读取的图片只能来自工作区或 Agent 上传目录。"""
    upload_raw = str(load_config().get("agent_upload_dir") or "~/.akm/cache").strip()
    upload_root = Path(upload_raw).expanduser().resolve()
    if _path_is_under(path, upload_root):
        return True
    workspace_root = _workspace_root()
    return workspace_root is not None and _path_is_under(path, workspace_root)


def _safe_resolve_workspace_path(raw: str, *, must_exist: bool = True) -> Path:
    """把模型传入的路径解析为工作区内的绝对路径，越界或缺失时抛 ValueError。

    安全约定：
    - 绝对路径必须位于工作区根目录内，否则拒绝；
    - 相对路径按工作区根目录拼接后再校验，`..` 等穿越写法会被拦截；
    - 最终用 resolve() 规范化，防止软链接把目标引到工作区之外；
    - must_exist 为 True 时目标必须存在（读写场景），False 时允许不存在
      （新建文件场景，但仍要求其父目录位于工作区内）。

    Returns:
        规范化后的绝对路径（Path 对象）。

    Raises:
        ValueError: 未配置工作区、路径越界或（must_exist 时）目标不存在。
    """
    root = _workspace_root()
    if root is None:
        raise ValueError("未配置 agent_workspace_root，工作区文件工具不可用")
    raw = str(raw or "").strip()
    if not raw:
        raise ValueError("path 不能为空")
    candidate = Path(raw)
    if not candidate.is_absolute():
        candidate = root / candidate
    candidate = candidate.resolve()
    try:
        candidate.relative_to(root)
    except ValueError:
        raise ValueError(f"路径 {raw} 超出工作区范围，禁止访问")
    if must_exist and not candidate.exists():
        raise ValueError(f"路径 {raw} 不存在")
    return candidate


async def _check_tool_enabled(flag: str, tool_name: str) -> None:
    """检查工具开关配置，未启用时抛出带提示的错误（由调用方捕获返回给模型）。"""
    if not load_config().get(flag):
        raise PermissionError(
            f"工具 {tool_name} 未启用：请在 config.json 中设置 {flag}=true"
        )


def _akm_api_base_url() -> str:
    """返回本机 AKM 服务地址，供工具以 HTTP 调用内部接口（如 markdown-kb）。

    与 akm/markdown_kb_hook.py 的约定保持一致：走 127.0.0.1 + server_port
    配置项，避免依赖外部可访问地址。
    """
    port = int(load_config().get("server_port", 8800) or 8800)
    return f"http://127.0.0.1:{port}"


def _image_default_model() -> str:
    """返回 config.json 中 image_supported_models 配置的首个模型。

    与 server.py 的 _default_image_generation_model 保持一致；为避免
    tools 与 server 互相 import 造成循环依赖，这里内联解析配置。
    """
    raw = str(load_config().get("image_supported_models") or "gpt-image-2").strip()
    models = [item.strip() for item in raw.split(",") if item.strip()]
    return (models or ["gpt-image-2"])[0]


def _image_request_timeout() -> float:
    """返回 config.json 中 image_request_timeout_sec 配置的图片请求超时秒数。"""
    try:
        timeout = float(load_config().get("image_request_timeout_sec", 300) or 300)
    except (TypeError, ValueError):
        timeout = 300.0
    return max(30.0, timeout)


def _read_image_file(path: str) -> tuple[str, bytes, str]:
    """读取本地图片文件，返回 (文件名, 字节内容, content_type)。

    路径只能位于当前工作区或 agent_upload_dir。若传入路径不存在，则提取
    文件名回退到 agent_upload_dir（默认 ~/.akm/cache）目录下查找。这样即使模型把
    http_url 中的 /agent-uploads/ 前缀误当成本地路径（例如编造出
    /data/agent-uploads/xxx.png），只要文件名一致仍能命中真实落盘文件。
    回退查找仅取文件名（basename），不会拼接目录，防止路径穿越。
    """
    file_path = Path(str(path or "")).expanduser().resolve()
    if not file_path.is_file():
        raw = str(load_config().get("agent_upload_dir") or "~/.akm/cache").strip()
        upload_dir = Path(raw).expanduser()
        fallback = (upload_dir / file_path.name).resolve()
        if fallback.is_file():
            file_path = fallback
        else:
            raise FileNotFoundError(f"图片文件不存在: {file_path}")
    # 直接传入的路径必须经过真实路径校验，避免借图片编辑能力读取并上传
    # 工作区或 Agent 上传目录之外的本机文件。
    if _request_workspace_root.get() is not None and not _allowed_image_path(file_path):
        raise ValueError("图片路径必须位于工作区或 agent_upload_dir 内")
    try:
        file_size = file_path.stat().st_size
    except OSError as exc:
        raise OSError(f"读取图片文件信息失败: {exc}")
    if file_size > _AGENT_IMAGE_MAX_INPUT_BYTES:
        raise ValueError(
            f"图片文件超过 {_AGENT_IMAGE_MAX_INPUT_BYTES // (1024 * 1024)}MB 限制"
        )
    content = file_path.read_bytes()
    content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
    return file_path.name, content, content_type


def _decode_image_base64(raw: str) -> tuple[str, bytes, str]:
    """解析图片 base64 数据，返回 (文件名, 字节内容, content_type)。

    兼容两种形态：带 ``data:image/png;base64,xxx`` 前缀的 data URL
    （Agent 对话中 image_url 内容块的形态），或去掉前缀的裸 base64。
    无前缀时按 image/png 处理；文件名用随机名 + 按 content_type 推断的
    扩展名，避免直接信任外部文件名。
    """
    text = str(raw or "").strip()
    if not text:
        raise ValueError("图片 base64 数据为空")
    content_type = "image/png"
    if text.startswith("data:"):
        header, sep, payload = text.partition(",")
        if not sep:
            raise ValueError("图片 data URL 格式非法，缺少逗号分隔的 base64 数据")
        mime = header[5:].split(";", 1)[0].strip()
        if mime and mime.lower().startswith("image/"):
            content_type = mime
        text = payload.strip()
    # base64 的编码尺寸最多约为原始字节数的 4/3；先按字符长度拒绝，
    # 防止 b64decode 为明显超限的输入分配大量内存。
    max_encoded_length = ((_AGENT_IMAGE_MAX_INPUT_BYTES + 2) // 3) * 4
    if len(text) > max_encoded_length:
        raise ValueError(
            f"图片 base64 数据超过 {_AGENT_IMAGE_MAX_INPUT_BYTES // (1024 * 1024)}MB 限制"
        )
    try:
        data = base64.b64decode(text, validate=True)
    except Exception:
        raise ValueError("图片 base64 解码失败，请检查数据是否完整")
    if not data:
        raise ValueError("图片 base64 数据为空")
    if len(data) > _AGENT_IMAGE_MAX_INPUT_BYTES:
        raise ValueError(
            f"图片 base64 数据超过 {_AGENT_IMAGE_MAX_INPUT_BYTES // (1024 * 1024)}MB 限制"
        )
    ext = mimetypes.guess_extension(content_type) or ".png"
    filename = f"upload-{uuid.uuid4().hex}{ext}"
    return filename, data, content_type


def _resolve_image(image_path: str = "", image_base64: str = "") -> tuple[str, bytes, str]:
    """按 image_path 或 image_base64 解析图片，返回 (文件名, 字节内容, content_type)。

    image_base64 非空时优先使用 base64 数据（适用于本地无文件的云端场景，
    AI 可直接把对话中的 data URL 传入）；否则回退读取 image_path 本地文件。
    两者都为空时抛出 ValueError。
    """
    if str(image_base64 or "").strip():
        return _decode_image_base64(image_base64)
    if str(image_path or "").strip():
        return _read_image_file(image_path)
    raise ValueError("image_path 与 image_base64 必须至少提供一个")


def _persist_image(data: bytes, content_type: str, source: str = "") -> tuple[str, str, str]:
    """把图片字节保存到 agent_upload_dir，返回 (文件名, 本地路径, HTTP URL)。

    content_type 为空时尝试从来源 URL 后缀推断；扩展名取 mimetypes 映射，
    无匹配时回退 .bin。HTTP URL 指向 /agent-uploads/{filename}，由服务端
    按 agent_upload_dir 提供访问，供前端或外部工具直接拉取。
    """
    raw = str(load_config().get("agent_upload_dir") or "~/.akm/cache").strip()
    upload_dir = Path(raw).expanduser()
    upload_dir.mkdir(mode=0o700, parents=True, exist_ok=True)
    if not content_type:
        content_type = mimetypes.guess_type(source.split("?", 1)[0])[0] or ""
    ext = mimetypes.guess_extension(content_type or "") or ".bin"
    if not ext.startswith("."):
        ext = f".{ext}"
    filename = f"{uuid.uuid4().hex}{ext}"
    local_path = str(upload_dir / filename)
    with open(local_path, "wb") as fh:
        fh.write(data)
    port = int(load_config().get("server_port", 8800) or 8800)
    http_url = f"http://127.0.0.1:{port}/agent-uploads/{filename}"
    return filename, local_path, http_url


async def _save_generated_image(item: dict[str, Any], client) -> None:
    """把生成/编辑结果的图片保存到上传目录，并填充 local_path/http_url。

    有 url 时通过连接池 client 下载；只有 b64_json 时直接解码。保存成功
    附加 local_path 与 http_url 字段，失败附加 save_error，均不影响主结果。
    """
    try:
        url = item.get("url")
        if url:
            resp = await client.get(url)
            resp.raise_for_status()
            data = resp.content
            content_type = (resp.headers or {}).get("content-type", "")
        else:
            b64 = item.pop("b64_json", None)
            if not b64:
                item["save_error"] = "无可用图片数据"
                return
            data = base64.b64decode(b64)
            content_type = ""
        _, local_path, http_url = _persist_image(data, content_type, url or "")
        item["local_path"] = local_path
        item["http_url"] = http_url
    except Exception as exc:
        logger.warning("[AgentTool] 保存生成图片失败: %s", exc)
        item["save_error"] = str(exc)


def _read_file_lines_sync(target: Path, start_line: int, max_lines: int) -> tuple[list[str], int, bool]:
    """同步迭代读取文件，仅保留目标行区间，返回 (选中的行, 总行数, 是否截断)。

    与旧实现的区别：不再用 readlines() 把整个文件读进内存，而是逐行迭代，
    只缓存目标区间的行；累计字节超过返回上限即停止（truncated=True）。
    运行在 asyncio.to_thread 中，避免阻塞事件循环。
    """
    selected: list[str] = []
    truncated = False
    total_lines = 0
    read_bytes = 0
    with open(target, "r", encoding="utf-8", errors="replace") as fh:
        for idx, line in enumerate(fh):
            total_lines = idx + 1
            if idx < start_line:
                continue
            if max_lines >= 0 and len(selected) >= max_lines:
                # 行数已达上限：继续计数总行数，但不再缓存内容
                continue
            read_bytes += len(line)
            if read_bytes > _WORKSPACE_READ_MAX_BYTES:
                truncated = True
                break
            selected.append(line)
    return selected, total_lines, truncated


async def _read_file_tool(
    path: str,
    offset: int = 0,
    limit: int = -1,
) -> str:
    """读取工作区内的文本文件，返回内容摘要（含长度限制与分页）。

    仅允许读取工作区根目录内的文件（见 _safe_resolve_workspace_path）。
    offset/limit 为行级分页：offset 从 0 开始，limit 为返回的最大行数，
    -1 表示读到文件结尾。单次返回字节数上限为 _WORKSPACE_READ_MAX_BYTES，
    超长时在结果中标记 truncated。
    """
    try:
        target = _safe_resolve_workspace_path(path, must_exist=True)
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)
    if not target.is_file():
        return json.dumps({"error": f"{path} 不是文件"}, ensure_ascii=False)
    try:
        if target.stat().st_size > _WORKSPACE_READ_MAX_FILE_BYTES:
            return json.dumps({
                "error": f"文件过大（{target.stat().st_size} 字节），超过单次读取上限 "
                         f"{_WORKSPACE_READ_MAX_FILE_BYTES}。可改用 akm_grep 定位关键内容，"
                         "或用 offset/limit 分页读取较小的文件。",
            }, ensure_ascii=False)
    except OSError as exc:
        return json.dumps({"error": f"读取文件信息失败: {exc}"}, ensure_ascii=False)

    offset = max(0, int(offset or 0))
    limit = int(limit if limit is not None else -1)
    try:
        # 同步 IO 放到线程池执行，避免阻塞事件循环
        lines, total_lines, truncated = await asyncio.to_thread(
            _read_file_lines_sync, target, offset, limit
        )
    except OSError as exc:
        return json.dumps({"error": f"读取文件失败: {exc}"}, ensure_ascii=False)

    text = "".join(lines)
    return json.dumps({
        "path": str(target),
        "start_line": offset,
        "total_lines": total_lines,
        "truncated": truncated,
        "content": text,
    }, ensure_ascii=False)


async def _list_dir_tool(path: str = "") -> str:
    """列出工作区内目录的条目（名称、类型、大小），供模型感知工作区结构。"""
    raw_path = str(path or "").strip()
    if not raw_path:
        raw_path = "."
    try:
        target = _safe_resolve_workspace_path(raw_path, must_exist=True)
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)
    if not target.is_dir():
        return json.dumps({"error": f"{path or '.'} 不是目录"}, ensure_ascii=False)
    try:
        entries = []
        total = 0
        for item in sorted(target.iterdir(), key=lambda p: p.name):
            total += 1
            if len(entries) >= _WORKSPACE_DIR_MAX_RESULTS:
                continue
            if item.is_dir():
                kind = "dir"
                size = None
            elif item.is_file():
                kind = "file"
                try:
                    size = item.stat().st_size
                except OSError:
                    size = None
            else:
                kind = "other"
                size = None
            entries.append({"name": item.name, "type": kind, "size": size})
    except OSError as exc:
        return json.dumps({"error": f"读取目录失败: {exc}"}, ensure_ascii=False)
    return json.dumps({"path": str(target), "entries": entries, "total": total}, ensure_ascii=False)


async def _glob_tool(pattern: str = "") -> str:
    """在工作区内按 glob 模式匹配文件路径（相对工作区根目录返回）。

    模式如 ``**/*.py``、``src/**/*.ts``。绝对模式或以 ``../`` 开头
    试图离开工作区的模式会被拒绝。
    """
    pattern = str(pattern or "").strip()
    if not pattern:
        return json.dumps({"error": "pattern 不能为空"}, ensure_ascii=False)
    if pattern.startswith("/") or pattern.startswith("../") or ".." in pattern.split("/"):
        return json.dumps({"error": "模式不允许离开工作区范围"}, ensure_ascii=False)
    root = _workspace_root()
    if root is None:
        return json.dumps({"error": "未配置 agent_workspace_root，工作区文件工具不可用"}, ensure_ascii=False)
    matches = []
    total = 0
    try:
        for p in root.glob(pattern):
            if not (p.is_file() or p.is_dir()):
                continue
            # 与 grep 一致：逐项 resolve 后校验仍位于工作区内，
            # 防止工作区内的软链接把匹配结果指向外部路径。
            try:
                if not _path_is_under(p.resolve(), root):
                    continue
            except OSError:
                continue
            total += 1
            if len(matches) < _WORKSPACE_GREP_MAX_RESULTS:
                matches.append(str(p.relative_to(root)))
    except (OSError, ValueError) as exc:
        return json.dumps({"error": f"匹配失败: {exc}"}, ensure_ascii=False)
    matches.sort()
    return json.dumps({"pattern": pattern, "matches": matches, "total": total}, ensure_ascii=False)


async def _grep_tool(
    pattern: str = "",
    path: str = "",
    case_sensitive: bool = False,
) -> str:
    """在工作区内按正则搜索文件内容，返回命中文件与行号。

    默认递归搜索整个工作区；path 指定时仅搜索该目录/文件（仍限工作区内）。
    结果条数上限为 _WORKSPACE_GREP_MAX_RESULTS。
    """
    pattern = str(pattern or "").strip()
    if not pattern:
        return json.dumps({"error": "pattern 不能为空"}, ensure_ascii=False)
    try:
        flags = 0 if case_sensitive else re.IGNORECASE
        rx = re.compile(pattern, flags)
    except re.error as exc:
        return json.dumps({"error": f"正则表达式非法: {exc}"}, ensure_ascii=False)
    raw_path = str(path or "").strip()
    if not raw_path:
        raw_path = "."
    try:
        root = _safe_resolve_workspace_path(raw_path, must_exist=True)
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)
    if not root.is_dir():
        return json.dumps({"error": f"{path or '.'} 不是目录"}, ensure_ascii=False)

    results = []
    scanned_bytes = 0
    try:
        for p in root.rglob("*"):
            if not p.is_file():
                continue
            # rglob 会返回工作区内指向外部文件的软链接，必须逐项 resolve
            # 后再次校验，不能只校验 grep 的起始目录。
            try:
                if not _path_is_under(p.resolve(), root):
                    continue
            except OSError:
                continue
            # 跳过隐藏目录与常见二进制/版本库目录，避免命中无意义内容
            rel = p.relative_to(root)
            parts = rel.parts
            if any(part.startswith(".") or part in ("node_modules", ".git", "__pycache__", "venv", ".venv", "build", "dist") for part in parts):
                continue
            # 资源预算：超大文件跳过、全局扫描字节预算封顶，防止全盘递归
            # 逐行正则匹配拖垮进程（该循环在事件循环内同步执行）。
            try:
                size = p.stat().st_size
            except OSError:
                continue
            if size > _WORKSPACE_GREP_MAX_FILE_BYTES:
                continue
            scanned_bytes += size
            if scanned_bytes > _WORKSPACE_GREP_MAX_SCAN_BYTES:
                break
            try:
                with open(p, "r", encoding="utf-8", errors="ignore") as fh:
                    for lineno, line in enumerate(fh, start=1):
                        # 跳过超长单行，避免灾难性正则（ReDoS）在超大输入上回退爆炸
                        if len(line) > _WORKSPACE_GREP_MAX_LINE_BYTES:
                            continue
                        if rx.search(line):
                            results.append({
                                "file": str(rel),
                                "line": lineno,
                                "content": line.rstrip("\n").encode("utf-8")[:_WORKSPACE_READ_MAX_BYTES].decode("utf-8", errors="ignore"),
                            })
                            break
            except OSError:
                continue
            if len(results) >= _WORKSPACE_GREP_MAX_RESULTS:
                break
    except OSError as exc:
        return json.dumps({"error": f"搜索失败: {exc}"}, ensure_ascii=False)
    return json.dumps({"pattern": pattern, "results": results, "total": len(results)}, ensure_ascii=False)


async def _file_info_tool(path: str) -> str:
    """返回工作区内文件或目录的元信息（类型、大小、修改时间）。"""
    try:
        target = _safe_resolve_workspace_path(path, must_exist=True)
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)
    try:
        stat = target.stat()
        kind = "dir" if target.is_dir() else "file"
        return json.dumps({
            "path": str(target),
            "type": kind,
            "size": stat.st_size,
            "mtime": datetime.datetime.fromtimestamp(stat.st_mtime).isoformat(),
        }, ensure_ascii=False)
    except OSError as exc:
        return json.dumps({"error": f"读取文件信息失败: {exc}"}, ensure_ascii=False)


def _atomic_write_text(target: Path, text: str) -> None:
    """以「临时文件 + os.replace」原子写入文本，避免进程中断/磁盘满时留下半个文件。

    临时文件创建在与目标同一目录，保证 os.replace 在同一文件系统内是原子操作。
    完成后按目标原权限（或默认 0644）修正临时文件权限，避免 mkstemp 的 0600
    覆盖掉原本可被同机其他用户读写的文件。
    """
    data = text.encode("utf-8")
    fd, tmp_path = tempfile.mkstemp(dir=str(target.parent), prefix=".akm-write-", suffix=".tmp")
    try:
        with os.fdopen(fd, "wb") as fh:
            fh.write(data)
        try:
            mode = target.stat().st_mode & 0o777
        except OSError:
            mode = 0o644
        os.chmod(tmp_path, mode)
        os.replace(tmp_path, target)
    except BaseException:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise


async def _write_file_tool(
    path: str,
    content: str,
    mode: str = "overwrite",
) -> str:
    """写文件：新建或覆盖工作区内的文本文件。仅 agent_write_tools_enabled=true 时可用。"""
    await _check_tool_enabled("agent_write_tools_enabled", "akm_write_file")
    mode = str(mode or "overwrite").strip()
    if mode not in ("overwrite", "append"):
        return json.dumps({"error": "mode 只能是 overwrite 或 append"}, ensure_ascii=False)
    payload = str(content or "")
    payload_bytes = len(payload.encode("utf-8"))
    if payload_bytes > _WORKSPACE_WRITE_MAX_BYTES:
        return json.dumps({
            "error": f"内容过大（{payload_bytes} 字节），超过单次写入上限 {_WORKSPACE_WRITE_MAX_BYTES}",
        }, ensure_ascii=False)
    try:
        target = _safe_resolve_workspace_path(path, must_exist=(mode == "append"))
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)
    if target.exists() and target.is_dir():
        return json.dumps({"error": f"{path} 是目录，不能写入"}, ensure_ascii=False)
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        if mode == "append":
            # 追加无原子语义，直接以 O_APPEND 打开续写
            with open(target, "a", encoding="utf-8") as fh:
                fh.write(payload)
        else:
            # 覆盖走原子替换，避免中断留下半个文件
            _atomic_write_text(target, payload)
    except OSError as exc:
        return json.dumps({"error": f"写入文件失败: {exc}"}, ensure_ascii=False)
    return json.dumps({"ok": True, "path": str(target), "mode": mode, "bytes_written": payload_bytes}, ensure_ascii=False)


async def _edit_file_tool(
    path: str,
    old_string: str = "",
    new_string: str = "",
    replace_all: bool = False,
    start_line: int = 0,
    end_line: int = 0,
    new_content: str = "",
) -> str:
    """结构化编辑工作区内的文本文件。仅 agent_write_tools_enabled=true 时可用。

    支持两种定位方式：
    - 行号模式：传入 start_line（1-based），把 [start_line, end_line] 行区间
      整体替换为 new_content；同时传 old_string 时作为锚点校验（确认原文
      确实位于目标行区间内），防止模型行号漂移后改错位置。推荐在读完文件、
      拿到精确行号后使用，比长字符串匹配更可靠。
    - 内容模式（默认，兼容旧行为）：将 old_string 替换为 new_string，
      replace_all 控制是否替换全部匹配。

    返回 old_lines / new_lines 或 replaced，便于模型感知编辑后的行数变化。
    """
    await _check_tool_enabled("agent_write_tools_enabled", "akm_edit_file")
    try:
        target = _safe_resolve_workspace_path(path, must_exist=True)
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)
    if not target.is_file():
        return json.dumps({"error": f"{path} 不是文件"}, ensure_ascii=False)
    try:
        content = target.read_text(encoding="utf-8", errors="replace")
    except OSError as exc:
        return json.dumps({"error": f"读取文件失败: {exc}"}, ensure_ascii=False)

    # ── 行号模式：基于行区间的结构化编辑 ──
    start_line = int(start_line or 0)
    end_line = int(end_line or 0)
    if start_line > 0:
        if start_line < 1:
            return json.dumps({"error": "start_line 必须 >= 1"}, ensure_ascii=False)
        if end_line and end_line < start_line:
            return json.dumps({"error": "end_line 不能小于 start_line"}, ensure_ascii=False)
        lines = content.splitlines()
        effective_end = end_line or start_line
        if effective_end > len(lines):
            return json.dumps(
                {"error": f"行号越界：文件共 {len(lines)} 行，请求编辑到第 {effective_end} 行"},
                ensure_ascii=False,
            )
        # 锚点校验：若同时提供 old_string，确认其恰好位于目标行区间内，
        # 防止模型基于旧快照的行号漂移后改错位置
        old_string = str(old_string or "")
        if old_string:
            target_slice = "\n".join(lines[start_line - 1:effective_end])
            if old_string not in target_slice:
                return json.dumps(
                    {"error": f"锚点校验失败：old_string 未在第 {start_line}-{effective_end} 行区间内找到"},
                    ensure_ascii=False,
                )
        new_lines = lines[: start_line - 1] + str(new_content or "").splitlines() + lines[effective_end:]
        new_text = "\n".join(new_lines)
        # 保留原文件尾部换行风格，避免无谓的全文件 diff
        if content.endswith("\n") and not new_text.endswith("\n"):
            new_text += "\n"
        if len(new_text.encode("utf-8")) > _WORKSPACE_WRITE_MAX_BYTES:
            return json.dumps({
                "error": f"编辑后文件过大（>{_WORKSPACE_WRITE_MAX_BYTES} 字节），拒绝写入",
            }, ensure_ascii=False)
        try:
            _atomic_write_text(target, new_text)
        except OSError as exc:
            return json.dumps({"error": f"写入文件失败: {exc}"}, ensure_ascii=False)
        return json.dumps({
            "ok": True,
            "path": str(target),
            "old_lines": effective_end - start_line + 1,
            "new_lines": len(str(new_content or "").splitlines()),
        }, ensure_ascii=False)

    # ── 内容模式：旧字符串替换（兼容原有行为）──
    old_string = str(old_string or "")
    if not old_string:
        return json.dumps({"error": "old_string 不能为空"}, ensure_ascii=False)
    count = content.count(old_string)
    if count == 0:
        return json.dumps({"error": "old_string 未在文件中找到"}, ensure_ascii=False)
    if not replace_all:
        count = 1
    new_text = content.replace(old_string, str(new_string or ""), count)
    if len(new_text.encode("utf-8")) > _WORKSPACE_WRITE_MAX_BYTES:
        return json.dumps({
            "error": f"编辑后文件过大（>{_WORKSPACE_WRITE_MAX_BYTES} 字节），拒绝写入",
        }, ensure_ascii=False)
    try:
        _atomic_write_text(target, new_text)
    except OSError as exc:
        return json.dumps({"error": f"写入文件失败: {exc}"}, ensure_ascii=False)
    return json.dumps({"ok": True, "path": str(target), "replaced": count}, ensure_ascii=False)


async def _make_dir_tool(path: str) -> str:
    """在工作区内创建目录（含父目录）。仅 agent_write_tools_enabled=true 时可用。"""
    await _check_tool_enabled("agent_write_tools_enabled", "akm_make_dir")
    try:
        target = _safe_resolve_workspace_path(path, must_exist=False)
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)
    try:
        target.mkdir(parents=True, exist_ok=True)
    except OSError as exc:
        return json.dumps({"error": f"创建目录失败: {exc}"}, ensure_ascii=False)
    return json.dumps({"ok": True, "path": str(target)}, ensure_ascii=False)


async def _delete_tool(path: str, recursive: bool = False) -> str:
    """删除工作区内的文件或目录。仅 agent_write_tools_enabled=true 时可用。

    默认只删除单个文件；``recursive=true`` 时允许删除目录并递归清除其中
    所有内容（用于批量删除场景）。无论是否 recursive，都拒绝删除工作区
    根目录本身。
    """
    await _check_tool_enabled("agent_write_tools_enabled", "akm_delete_file")
    try:
        target = _safe_resolve_workspace_path(path, must_exist=True)
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)
    root = _workspace_root()
    if target == root:
        return json.dumps({"error": "禁止删除工作区根目录"}, ensure_ascii=False)
    if target.is_dir() and not recursive:
        return json.dumps(
            {"error": "目标是一个目录，批量删除需设置 recursive=true；请确认后重试"},
            ensure_ascii=False,
        )
    try:
        if target.is_dir():
            shutil.rmtree(target)
        else:
            target.unlink()
    except OSError as exc:
        return json.dumps({"error": f"删除失败: {exc}"}, ensure_ascii=False)
    return json.dumps({"ok": True, "path": str(target)}, ensure_ascii=False)


async def _run_workspace_argv(argv: list[str], root: Path, timeout: int, label: str) -> str:
    """在工作区执行已经由服务端构造的 argv，并统一限制时间与输出大小。"""
    timeout = max(1, min(int(timeout or _WORKSPACE_SHELL_TIMEOUT_SEC), 300))
    try:
        proc = await asyncio.create_subprocess_exec(
            *argv,
            cwd=str(root),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
        )
        try:
            output, _ = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        except asyncio.TimeoutError:
            proc.kill()
            await proc.wait()
            return json.dumps({"error": f"{label}执行超时（{timeout} 秒），已终止"}, ensure_ascii=False)
        text = output.decode("utf-8", errors="replace")
        encoded = text.encode("utf-8")
        truncated = len(encoded) > _WORKSPACE_READ_MAX_BYTES
        if truncated:
            text = encoded[:_WORKSPACE_READ_MAX_BYTES].decode("utf-8", errors="ignore")
        return json.dumps({"exit_code": proc.returncode, "truncated": truncated, "output": text}, ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"error": f"执行{label}失败: {exc}"}, ensure_ascii=False)


def _seatbelt_profile(root: Path) -> str:
    """构造 macOS seatbelt 黑名单式 sandbox profile（SBPL）。

    策略：全局禁写 + 拒绝敏感路径 + 拒绝家目录元数据列举，然后**后置**
    allow 放行工作区与临时目录。seatbelt 规则按声明顺序求值、后面的规则
    覆盖前面的（last-match-wins），因此 allow 必须写在所有 deny 之后，
    否则 deny 敏感目录（如 ~/Desktop）会把位于其下的工作区一并锁死。
    这是「限制敏感读写」级别的隔离，不是真正的 chroot，命令仍可读系统
    目录、访问网络；但能挡住读密钥/配置文件/命令历史/家目录 dotfile、
    列家目录与 /var、/tmp 结构、往系统写文件的越界。
    """
    ws = str(root)
    home = str(Path.home())
    tmp = os.environ.get("TMPDIR") or "/tmp"
    sensitive = [
        f'"{home}/.ssh"',
        f'"{home}/.aws"',
        f'"{home}/.akm"',
        f'"{home}/Downloads"',
        f'"{home}/Documents"',
        f'"{home}/Desktop"',
        f'"{home}/Library"',
        # home 根目录点文件：shell 配置与命令历史、git/npm 凭据等
        f'"{home}/.zshrc"',
        f'"{home}/.zprofile"',
        f'"{home}/.bash_profile"',
        f'"{home}/.bashrc"',
        f'"{home}/.zsh_history"',
        f'"{home}/.bash_history"',
        f'"{home}/.gitconfig"',
        f'"{home}/.git-credentials"',
        f'"{home}/.npmrc"',
        f'"{home}/.netrc"',
        # 系统目录：/etc 是 /private/etc 的符号链接，cat /etc/passwd 实际打开
        # /private/etc/passwd；/var 与 /tmp 同理。deny 必须同时覆盖符号链接
        # 的两个形态才有效
        '"/etc"',
        '"/private/etc"',
        '"/var/log"',
        '"/private/var/log"',
        '"/var/db"',
        '"/private/var/db"',
        '"/tmp"',
        '"/private/tmp"',
    ]
    deny_read = " ".join(f'(subpath {p})' for p in sensitive)
    allow_write = (
        f'(subpath "{ws}") (subpath "{tmp}")'
        ' (literal "/dev/null") (literal "/dev/urandom") (literal "/dev/zero")'
    )
    allow_read = (
        f'(subpath "{ws}") (subpath "{tmp}")'
        ' (literal "/dev/null") (literal "/dev/urandom") (literal "/dev/zero")'
    )
    return (
        "(version 1)\n"
        "(allow default)\n"
        "(deny file-write*)\n"
        f"(deny file-read* {deny_read})\n"
        # 家目录的整体元数据列举（如 ls ~）也会泄露目录名，单独拒绝；
        # 后面的 allow 会为工作区/临时目录重新放行对应 metadata 访问。
        f'(deny file-read-metadata (subpath "{home}"))\n'
        f"(allow file-write* {allow_write})\n"
        f"(allow file-read* {allow_read})\n"
        f'(allow file-read-metadata (subpath "{ws}") (subpath "{tmp}"))\n'
    )


_seatbelt_lib = None
_seatbelt_api_ok = None


def _seatbelt_available() -> bool:
    """探测当前运行环境是否提供 libSystem 的 sandbox_init_with_parameters。

    结果缓存在模块级变量，避免每次执行 shell 都重复探测。py2app 打包后
    libSystem.dylib 始终存在；若 API 缺失（如非 macOS 或未来系统移除），
    返回 False，调用方应退回无隔离执行并记录警告。
    """
    global _seatbelt_lib, _seatbelt_api_ok
    if _seatbelt_api_ok is not None:
        return _seatbelt_api_ok
    try:
        _seatbelt_lib = ctypes.CDLL("libSystem.dylib")
        getattr(_seatbelt_lib, "sandbox_init_with_parameters")
        _seatbelt_api_ok = True
    except (AttributeError, OSError):
        _seatbelt_lib = None
        _seatbelt_api_ok = False
    return _seatbelt_api_ok


def _seatbelt_preexec(profile: str):
    """构造 preexec_fn：在 fork 出的子进程、exec 之前应用 seatbelt 沙箱。

    只对即将 exec 的 shell 子进程生效，AKM 主进程不受影响。用
    sandbox_init_with_parameters 应用黑名单式 profile；失败时直接 os._exit
    避免子进程带沙箱进入 exec，父进程的 wait_for 会捕获非零返回码。
    """
    def _apply():
        err = ctypes.c_char_p()
        lib: ctypes.CDLL = _seatbelt_lib  # type: ignore[assignment]
        rc = lib.sandbox_init_with_parameters(
            profile.encode("utf-8"), 0, None, ctypes.byref(err)
        )
        if rc != 0:
            err_text = (err.value or b"").decode("utf-8", "replace")
            msg = f"SANDBOX_INIT_FAILED rc={rc} {err_text}\n"
            os.write(2, msg.encode("utf-8", "replace"))
            os._exit(125)
    return _apply


async def _run_workspace_shell(command: str, root: Path, timeout: int, label: str) -> str:
    """在工作区用系统 shell 执行命令字符串，并统一限制时间与输出大小。

    shell 语义（管道 / 通配符 / 重定向等）由系统 shell 解释，模型可编写
    任意命令；cwd 固定为工作区根目录。执行受超时与输出字节上限约束。
    当配置 agent_run_shell_sandbox=true 且本机提供 sandbox_init_with_parameters
    时，用 seatbelt 沙箱隔离 shell 子进程（见 _seatbelt_profile）。
    """
    timeout = max(1, min(int(timeout or _WORKSPACE_SHELL_TIMEOUT_SEC), 300))
    sandbox = bool(load_config().get("agent_run_shell_sandbox", False))
    preexec_fn = None
    if sandbox:
        if _seatbelt_available():
            preexec_fn = _seatbelt_preexec(_seatbelt_profile(root))
        else:
            logger.warning("agent_run_shell_sandbox=true 但本机缺少 sandbox_init_with_parameters，退回无隔离执行")
    try:
        proc = await asyncio.create_subprocess_shell(
            command,
            cwd=str(root),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            preexec_fn=preexec_fn,
        )
        try:
            output, _ = await asyncio.wait_for(proc.communicate(), timeout=timeout)
        except asyncio.TimeoutError:
            proc.kill()
            await proc.wait()
            return json.dumps({"error": f"{label}执行超时（{timeout} 秒），已终止"}, ensure_ascii=False)
        text = output.decode("utf-8", errors="replace")
        encoded = text.encode("utf-8")
        truncated = len(encoded) > _WORKSPACE_READ_MAX_BYTES
        if truncated:
            text = encoded[:_WORKSPACE_READ_MAX_BYTES].decode("utf-8", errors="ignore")
        return json.dumps({"exit_code": proc.returncode, "truncated": truncated, "output": text}, ensure_ascii=False)
    except Exception as exc:
        return json.dumps({"error": f"执行{label}失败: {exc}"}, ensure_ascii=False)


async def _run_shell_tool(command: str, timeout: int = 0) -> str:
    """在工作区用 shell 执行命令字符串并返回 stdout+stderr。

    命令由模型/客户端直接传入，系统 shell 解释执行（支持管道、通配符、
    重定向等），cwd 固定为工作区根目录；超时与输出大小统一受限。
    这是显式开启的主机级进程执行能力，仅 agent_run_shell_enabled=true
    时注册，管理员应通过 tool_policy_guard 等插件策略约束调用。
    """
    await _check_tool_enabled("agent_run_shell_enabled", "akm_run_shell")
    command = str(command or "").strip()
    if not command:
        return json.dumps({"error": "command 不能为空"}, ensure_ascii=False)
    root = _workspace_root()
    if root is None:
        return json.dumps({"error": "未配置 agent_workspace_root，工作区文件工具不可用"}, ensure_ascii=False)
    return await _run_workspace_shell(command, root, timeout, "shell 命令")


def _xlsx_range_bounds(range_ref: str) -> tuple | None:
    """把 Excel 区间字符串（如 "B2:C6"）解析为 (min_col, min_row, max_col, max_row)。

    非法格式返回 None，由调用方统一报错。
    """
    try:
        from openpyxl.utils.cell import range_boundaries
        bounds = range_boundaries(str(range_ref).replace("$", "").upper())
        if bounds is None or bounds[0] is None or bounds[2] is None:
            return None
        return bounds
    except (ImportError, ValueError):
        return None


def _xlsx_reference(ws, range_ref: str):
    """按 Excel 区间字符串构造 openpyxl Reference（供图表使用）。"""
    from openpyxl.chart import Reference
    bounds = _xlsx_range_bounds(range_ref)
    if bounds is None:
        raise ValueError(f"非法单元格区间: {range_ref}")
    min_col, min_row, max_col, max_row = bounds
    return Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)


def _apply_xlsx_styles(wb, styles: Any) -> str | None:
    """把 styles 列表应用到工作簿，返回错误信息（无错误返回 None）。

    styles 每项: {sheet?, cell, bold?, italic?, size?, color?, fill?,
    align?, number_format?}。color/fill 为不带 # 的十六进制色值。
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    if not isinstance(styles, list):
        return "styles 必须是数组"
    for item in styles:
        if not isinstance(item, dict):
            return "styles 中的每一项必须是对象"
        cell_ref = str(item.get("cell") or "").strip().upper()
        if not cell_ref:
            return "styles 每一项都必须包含 cell（如 A1）"
        sheet_name = str(item.get("sheet") or "Sheet1")
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
        cell = ws[cell_ref]

        if any(k in item for k in ("bold", "italic", "size", "color")):
            cell.font = Font(
                bold=item.get("bold"),
                italic=item.get("italic"),
                size=item.get("size"),
                color=str(item.get("color") or "").strip("#") or None,
            )
        fill = str(item.get("fill") or "").strip("#")
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        align = str(item.get("align") or "").strip().lower()
        if align:
            cell.alignment = Alignment(horizontal=align) if align in ("left", "center", "right", "fill", "justify", "center_continuous", "distributed") else Alignment(vertical=align)
        if item.get("number_format"):
            cell.number_format = str(item["number_format"])
    return None


def _apply_xlsx_layout(wb, column_widths: Any, row_heights: Any, merge_cells: Any, freeze_panes: Any) -> str | None:
    """应用列宽/行高/合并单元格/冻结窗格，返回错误信息（无错误返回 None）。

    各参数均为 {sheet名: 配置} 映射；单 sheet 场景 sheet 键可省略，默认 Sheet1。
    """
    def resolve(entries: Any, default_sheet: str):
        if entries is None:
            return {}
        if isinstance(entries, dict):
            if "Sheet1" in entries or "sheet" in entries or any(str(k).startswith("Sheet") for k in entries):
                return {str(k): v for k, v in entries.items()}
            return {default_sheet: entries}
        return {default_sheet: entries}

    for sheet_name, config in resolve(column_widths, "Sheet1").items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
        if not isinstance(config, dict):
            return "column_widths 配置必须是 {列名: 宽度} 映射"
        for col, width in config.items():
            ws.column_dimensions[str(col).upper()].width = float(width)

    for sheet_name, config in resolve(row_heights, "Sheet1").items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
        if not isinstance(config, dict):
            return "row_heights 配置必须是 {行号: 高度} 映射"
        for row, height in config.items():
            ws.row_dimensions[int(row)].height = float(height)

    for sheet_name, config in resolve(merge_cells, "Sheet1").items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
        if isinstance(config, str):
            config = [config]
        if not isinstance(config, list):
            return "merge_cells 配置必须是区间字符串数组"
        for rng in config:
            ws.merge_cells(str(rng).upper())

    for sheet_name, config in resolve(freeze_panes, "Sheet1").items():
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
        if not isinstance(config, str):
            return "freeze_panes 配置必须是单元格引用字符串（如 A2）"
        ws.freeze_panes = str(config).upper()
    return None


def _apply_xlsx_charts(wb, charts: Any) -> str | None:
    """把图表定义列表应用到工作簿，返回错误信息（无错误返回 None）。

    charts 每项: {sheet?, type, title?, data_range, categories_range?,
    x_title?, y_title?, anchor?, legend?}。type 支持 bar/line/pie/
    scatter/area/doughnut。
    """
    from openpyxl.chart import AreaChart, BarChart, DoughnutChart, LineChart, PieChart, ScatterChart

    chart_types = {
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "area": AreaChart,
        "doughnut": DoughnutChart,
    }
    if not isinstance(charts, list):
        return "charts 必须是数组"
    for item in charts:
        if not isinstance(item, dict):
            return "charts 中的每一项必须是对象"
        ctype = str(item.get("type") or "").strip().lower()
        cls = chart_types.get(ctype)
        if cls is None:
            return f"不支持的图表类型: {ctype}（可用: {', '.join(chart_types)}）"
        data_range = str(item.get("data_range") or "").strip().upper()
        if not data_range:
            return "charts 每一项都必须包含 data_range（如 B2:B6）"
        sheet_name = str(item.get("sheet") or "Sheet1")
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

        chart = cls()
        if item.get("title"):
            chart.title = str(item["title"])
        chart.add_data(_xlsx_reference(ws, data_range), titles_from_data=bool(item.get("titles_from_data")))
        categories = str(item.get("categories_range") or "").strip().upper()
        if categories:
            chart.set_categories(_xlsx_reference(ws, categories))
        if item.get("x_title"):
            if getattr(chart, "x_axis", None) is not None:
                chart.x_axis.title = str(item["x_title"])
        if item.get("y_title"):
            if getattr(chart, "y_axis", None) is not None:
                chart.y_axis.title = str(item["y_title"])
        if item.get("legend") is False:
            chart.legend = None
        ws.add_chart(chart, str(item.get("anchor") or "F2").upper())
    return None


async def _xlsx_tool(
    action: str,
    path: str,
    data: Any = None,
    sheet: str = "",
    overwrite: bool = False,
    updates: Any = None,
    styles: Any = None,
    column_widths: Any = None,
    row_heights: Any = None,
    merge_cells: Any = None,
    freeze_panes: Any = None,
    charts: Any = None,
) -> str:
    """创建或修改工作区内的 .xlsx 电子表格文件（基于 openpyxl）。

    仅 agent_write_tools_enabled=true 时注册。action 区分两种操作：
    - ``create``：新建 xlsx。``data`` 为二维数组（每行一个子数组）或
      {sheet_name: 二维数组} 的映射；纯数组时写入 ``sheet`` 指定的工作表
      （默认 "Sheet1"）。目标已存在且 overwrite=false 时拒绝，防止误覆盖。
    - ``edit``：修改已有 xlsx。``updates`` 为 [{sheet, cell, value}] 列表，
      逐条写入指定单元格（cell 形如 "A1"，sheet 默认 "Sheet1"）。

    两种 action 共用以下可选自定义参数：
    - ``styles``：[{sheet?, cell, bold?, italic?, size?, color?, fill?,
      align?, number_format?}] 设置单元格字体/背景/对齐/数字格式。
    - ``column_widths`` / ``row_heights``：{sheet?: {列名/行号: 尺寸}} 映射，
      sheet 键可省略（默认 Sheet1）。
    - ``merge_cells``：{sheet?: [区间, ...]} 合并单元格。
    - ``freeze_panes``：{sheet?: "A2"} 冻结窗格。
    - ``charts``：[{sheet?, type, title?, data_range, categories_range?,
      x_title?, y_title?, anchor?, legend?}] 添加图表，type 支持 bar/line/
      pie/scatter/area/doughnut。
    公式支持：value 或 data 中以 "=" 开头的字符串会按公式写入（如
    "=SUM(B2:B6)"）。所有文件路径限定在工作区内，返回值是 JSON 字符串。
    """
    await _check_tool_enabled("agent_write_tools_enabled", "akm_xlsx")
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        return json.dumps({"error": "服务端未安装 openpyxl，无法使用 xlsx 工具"}, ensure_ascii=False)

    action = str(action or "").strip()
    if action not in ("create", "edit"):
        return json.dumps({"error": "action 只能是 create 或 edit"}, ensure_ascii=False)
    try:
        target = _safe_resolve_workspace_path(path, must_exist=(action == "edit"))
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)

    try:
        if action == "create":
            if target.exists() and not overwrite:
                return json.dumps({"error": f"{path} 已存在，如需覆盖请传 overwrite=true"}, ensure_ascii=False)
            if data is None:
                return json.dumps({"error": "create 需要 data 参数"}, ensure_ascii=False)
            wb = Workbook()
            active = wb.active
            if active is not None:
                wb.remove(active)  # 移除默认空白工作表，按 data 重建
            if isinstance(data, dict):
                sheets = {str(k): v for k, v in data.items()}
            else:
                sheets = {str(sheet or "Sheet1"): data}
            for name, rows in sheets.items():
                if not isinstance(rows, list):
                    return json.dumps({"error": f"工作表 {name} 的数据必须是二维数组"}, ensure_ascii=False)
                ws = wb.create_sheet(title=str(name))
                for row in rows:
                    if not isinstance(row, list):
                        return json.dumps({"error": f"工作表 {name} 的行必须是数组"}, ensure_ascii=False)
                    ws.append(row)
        else:  # action == "edit"
            if not updates:
                return json.dumps({"error": "edit 需要 updates 参数（[{sheet, cell, value}, ...]）"}, ensure_ascii=False)
            if not isinstance(updates, list):
                return json.dumps({"error": "updates 必须是数组"}, ensure_ascii=False)
            wb = load_workbook(target)
            for item in updates:
                if not isinstance(item, dict):
                    return json.dumps({"error": "updates 中的每一项必须是 {sheet, cell, value}"}, ensure_ascii=False)
                cell_ref = str(item.get("cell") or "").strip().upper()
                if not cell_ref:
                    return json.dumps({"error": "updates 每一项都必须包含 cell（如 A1）"}, ensure_ascii=False)
                sheet_name = str(item.get("sheet") or "Sheet1")
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)
                ws[cell_ref] = item.get("value")

        # 自定义参数：样式 / 布局 / 图表（create 与 edit 共用）
        if styles is not None:
            err = _apply_xlsx_styles(wb, styles)
            if err:
                return json.dumps({"error": err}, ensure_ascii=False)
        if any(v is not None for v in (column_widths, row_heights, merge_cells, freeze_panes)):
            err = _apply_xlsx_layout(wb, column_widths, row_heights, merge_cells, freeze_panes)
            if err:
                return json.dumps({"error": err}, ensure_ascii=False)
        if charts is not None:
            err = _apply_xlsx_charts(wb, charts)
            if err:
                return json.dumps({"error": err}, ensure_ascii=False)

        target.parent.mkdir(parents=True, exist_ok=True)
        wb.save(target)
    except KeyError as exc:
        return json.dumps({"error": f"修改文件失败: {exc}"}, ensure_ascii=False)
    except (ValueError, TypeError) as exc:
        return json.dumps({"error": f"写入文件失败: {exc}"}, ensure_ascii=False)
    except OSError as exc:
        return json.dumps({"error": f"保存文件失败: {exc}"}, ensure_ascii=False)
    return json.dumps(
        {"ok": True, "action": action, "path": str(target), **({"updated": len(updates)} if action == "edit" else {})},
        ensure_ascii=False,
    )


def _git_paths(paths: list[str] | None) -> list[str]:
    """规范化 git 路径参数，防止通过路径穿越让 git 操作工作区外文件。"""
    if paths is None:
        return []
    if not isinstance(paths, list) or len(paths) > 100:
        raise ValueError("paths 必须是最多 100 项的数组")
    normalized = []
    for raw in paths:
        path = str(raw or "").strip()
        candidate = Path(path)
        if not path or candidate.is_absolute() or ".." in candidate.parts:
            raise ValueError("paths 只能包含工作区内的相对路径")
        normalized.append(path)
    return normalized


async def _run_git_tool(
    operation: str,
    paths: list[str] | None = None,
    message: str = "",
    revision: str = "",
    staged: bool = False,
    limit: int = 20,
    timeout: int = 0,
) -> str:
    """执行固定集合的结构化 git 操作，不接受自由命令字符串。"""
    await _check_tool_enabled("agent_git_enabled", "akm_run_git")
    root = _workspace_root()
    if root is None:
        return json.dumps({"error": "未配置 agent_workspace_root，工作区文件工具不可用"}, ensure_ascii=False)
    operation = str(operation or "").strip()
    try:
        safe_paths = _git_paths(paths)
    except ValueError as exc:
        return json.dumps({"error": str(exc)}, ensure_ascii=False)

    if operation == "status":
        command = ["status", "--short"]
    elif operation == "diff":
        command = ["diff"] + (["--staged"] if staged else []) + (["--"] + safe_paths if safe_paths else [])
    elif operation == "log":
        command = ["log", f"-n{max(1, min(int(limit or 20), 100))}", "--oneline"]
    elif operation == "show":
        revision = str(revision or "HEAD").strip()
        if not re.fullmatch(r"[A-Za-z0-9._/@^~:-]+", revision):
            return json.dumps({"error": "revision 包含不允许的字符"}, ensure_ascii=False)
        command = ["show", revision]
    elif operation in {"add", "restore", "reset"}:
        if not safe_paths:
            return json.dumps({"error": f"{operation} 必须提供 paths"}, ensure_ascii=False)
        command = [operation, "--"] + safe_paths
    elif operation == "commit":
        message = str(message or "").strip()
        if not message or len(message) > 500:
            return json.dumps({"error": "commit 必须提供不超过 500 字符的 message"}, ensure_ascii=False)
        command = ["commit", "-m", message]
    elif operation == "branch":
        command = ["branch", "--show-current"]
    else:
        return json.dumps({"error": f"不支持的 git operation: {operation}"}, ensure_ascii=False)

    argv = [
        "git", "-c", "core.pager=cat", "-c", "color.ui=false",
        "-c", "core.hooksPath=/dev/null", "--no-pager",
    ] + command
    return await _run_workspace_argv(argv, root, timeout, f"git {operation}")


def build_builtin_tools(app: FastAPI) -> list[ToolDef]:
    """创建与当前服务实例绑定的只读调试工具。

    工具刻意只暴露排障所需的运行元数据。密钥明文、审计请求体、响应体和
    请求头都不会进入模型上下文，避免 Agent 调用意外扩大敏感数据暴露面。
    """

    def get_status() -> dict[str, Any]:
        """返回健康监护、审计队列和已加载插件的摘要状态。"""
        monitor = getattr(app.state, "health_monitor", None)
        audit_queue = getattr(app.state, "audit_log_queue", None)
        plugin_manager = getattr(app.state, "plugin_manager", None)
        plugins = getattr(plugin_manager, "plugins", {})
        return {
            "health": monitor.detail_payload() if monitor is not None else {},
            "audit_queue": {
                "size": audit_queue.qsize() if audit_queue is not None else 0,
                "maxsize": getattr(audit_queue, "maxsize", 0),
                "dropped_count": getattr(audit_queue, "dropped_count", 0),
                "failure_count": getattr(audit_queue, "failure_count", 0),
                "worker_alive": audit_queue.worker_alive() if audit_queue is not None else False,
            },
            "plugins": [
                {
                    "name": name,
                    "enabled": plugin.enabled,
                    "runtime_ready": plugin.runtime_ready,
                }
                for name, plugin in plugins.items()
            ],
        }

    def get_keys() -> list[dict[str, Any]]:
        """返回 Key 的非敏感连接与模型配置，绝不返回 API Key。"""
        return [
            {
                "alias": key.get("alias", ""),
                "provider": key.get("provider", ""),
                "models": key_model_list(key),
                "priority": key.get("priority", 0),
                "status": key.get("status", ""),
            }
            for key in list_keys()
        ]

    def get_keys_summary() -> dict[str, Any]:
        """返回当前 Key 总数与每个 Key 的模型清单，不返回任何密钥信息。"""
        keys = [
            {
                "alias": key.get("alias", ""),
                "provider": key.get("provider", ""),
                "models": key_model_list(key),
                "status": key.get("status", ""),
            }
            for key in list_keys()
        ]
        return {
            "total": len(keys),
            "keys": keys,
        }

    def get_time() -> dict[str, Any]:
        """返回服务器当前时间，含本地 ISO 时间、UTC 时间、UNIX 时间戳与时区。"""
        now = datetime.datetime.now().astimezone()
        return {
            "iso": now.isoformat(),
            "utc_iso": now.astimezone(datetime.timezone.utc).isoformat(),
            "unix": int(now.timestamp()),
            "timezone": str(now.tzinfo),
        }

    async def get_logs(
        limit: int = 20,
        status: str = "all",
        days: int = 1,
        key_alias: str = "",
    ) -> list[dict[str, Any]]:
        """返回近期审计元数据，不包含任意请求或响应内容。"""
        limit = max(1, min(int(limit), 50))
        days = max(0, min(int(days), 30))
        if status not in {"all", "success", "failed"}:
            raise ValueError("status 只能是 all、success 或 failed")
        logs = await list_logs_async(
            limit=limit,
            status=status,
            days=days,
            key_alias=str(key_alias or ""),
        )
        return [
            {
                "id": log.get("id"),
                "timestamp": log.get("timestamp", ""),
                "provider": log.get("provider", ""),
                "key_alias": log.get("key_alias", ""),
                "model": log.get("model", ""),
                "status_code": log.get("status_code", 0),
                "latency_ms": log.get("latency_ms", 0),
                "prompt_tokens": log.get("prompt_tokens", 0),
                "completion_tokens": log.get("completion_tokens", 0),
                "error": log.get("error", ""),
            }
            for log in logs
        ]

    def get_usage_stats(days: int = 0) -> dict[str, Any]:
        """查询 Token 用量统计（默认同时返回最近 1/7/30 天）。

        费用字段与模型单价表仅在 config.json 开启 ``cost_stats_enabled`` 时返回；
        估算值不能替代供应商账单。

        Args:
            days: 查询窗口天数。传 1 / 7 / 30 时只返回该窗口；
                传 0（默认）或其它值时返回 1、7、30 三个窗口。
        """
        # 延迟导入：tools 在 server 启动阶段注册，避免模块级循环依赖。
        from akm.cost_estimate import DEFAULT_PRICING_TABLE, pricing_snapshot
        from akm.server import _get_stats

        cfg = load_config()
        cost_enabled = bool(cfg.get("cost_stats_enabled", False))
        pricing_table = str(cfg.get("cost_pricing_table") or DEFAULT_PRICING_TABLE)
        # 只允许固定窗口，避免模型随意拉超长区间撑爆上下文。
        allowed_windows = (1, 7, 30)
        try:
            days_n = int(days or 0)
        except (TypeError, ValueError):
            days_n = 0
        if days_n in allowed_windows:
            window_days = (days_n,)
        else:
            window_days = allowed_windows

        def _summarize_bucket(bucket: dict[str, Any]) -> dict[str, Any]:
            """压缩 by_* 分桶：保留 token / 请求数，费用开启时再带 cost。"""
            out: dict[str, Any] = {
                "prompt": int(bucket.get("prompt", 0) or 0),
                "completion": int(bucket.get("completion", 0) or 0),
                "total": int(bucket.get("total", 0) or 0),
                "cached": int(bucket.get("cached", 0) or 0),
                "requests": int(bucket.get("requests", 0) or 0),
            }
            if cost_enabled and "cost" in bucket:
                out["cost"] = float(bucket.get("cost", 0) or 0)
                out["currency"] = str(bucket.get("currency") or "$")
            return out

        def _window_payload(raw: dict[str, Any], window: int) -> dict[str, Any]:
            """把 _get_stats 原始结果压成 Agent 友好摘要。"""
            payload: dict[str, Any] = {
                "days": window,
                "total_requests": int(raw.get("total_requests", 0) or 0),
                "total_prompt_tokens": int(raw.get("total_prompt_tokens", 0) or 0),
                "total_completion_tokens": int(raw.get("total_completion_tokens", 0) or 0),
                "total_tokens": int(raw.get("total_tokens", 0) or 0),
                "total_cached_tokens": int(raw.get("total_cached_tokens", 0) or 0),
                "by_model": {
                    str(name): _summarize_bucket(item)
                    for name, item in (raw.get("by_model") or {}).items()
                },
                "by_provider": {
                    str(name): _summarize_bucket(item)
                    for name, item in (raw.get("by_provider") or {}).items()
                },
                "by_key": {
                    str(name): _summarize_bucket(item)
                    for name, item in (raw.get("by_key") or {}).items()
                },
                "cached_at": raw.get("cached_at") or "",
            }
            if cost_enabled:
                payload["total_cost"] = float(raw.get("total_cost", 0) or 0)
                payload["cost_currency"] = str(raw.get("cost_currency") or "$")
                if raw.get("costs_by_currency") is not None:
                    payload["costs_by_currency"] = raw.get("costs_by_currency")
            return payload

        windows: dict[str, Any] = {}
        for window in window_days:
            windows[str(window)] = _window_payload(_get_stats(window), window)

        result: dict[str, Any] = {
            "windows": windows,
            "cost_stats_enabled": cost_enabled,
        }
        # 费用与单价表都只在 cost_stats_enabled 开启时返回，避免未开启时干扰上下文。
        if cost_enabled:
            result["pricing"] = pricing_snapshot(pricing_table)
            result["pricing_unit"] = "USD per 1M tokens (input / input_cache / output)"
            result["cost_note"] = "费用为本地单价表估算，不能替代供应商账单"
        else:
            result["cost_note"] = (
                "费用估算未开启：在 config.json 设置 cost_stats_enabled=true 后"
                "再查询可返回 total_cost 与模型单价表"
            )
        return result

    def get_config() -> dict[str, Any]:
        """返回 AKM 运行配置。密钥类字段（agent_api_token、tavily_api_key）
        不做明文透出：已配置时标记为"已配置"，避免敏感信息进入模型上下文。"""
        cfg = load_config()
        redacted: dict[str, Any] = {
            "agent_api_token": ("已配置" if cfg.get("agent_api_token") else ""),
            "tavily_api_key": ("已配置" if cfg.get("tavily_api_key") else ""),
            "agent_email_smtp_password": ("已配置" if cfg.get("agent_email_smtp_password") else ""),
        }
        safe = {key: value for key, value in cfg.items() if key not in redacted}
        safe.update(redacted)
        return safe

    def list_plugins() -> list[dict[str, Any]]:
        """返回已加载插件的非敏感摘要（名称、版本、分类、描述、启用状态与来源）。"""
        pm = getattr(app.state, "plugin_manager", None)
        if pm is None:
            return []
        try:
            items = pm.get_plugin_list()
        except Exception:
            logger.warning("[AgentTool] akm_list_plugins 读取插件列表失败", exc_info=True)
            return []
        return [
            {
                "name": item.get("name", ""),
                "version": item.get("version", ""),
                "category": item.get("category", ""),
                "description": item.get("description", ""),
                "builtin": item.get("builtin", False),
                "enabled": item.get("enabled", False),
                "source": item.get("source", ""),
            }
            for item in items
        ]

    def list_sessions() -> list[dict[str, Any]]:
        """列出历史 Agent 会话的元信息（不含消息正文），按更新时间倒序。"""
        from akm.agent_runtime.sessions import SessionStore
        return SessionStore().list()

    def load_session(name: str, limit: int = 20) -> dict[str, Any]:
        """读取历史 Agent 会话的最近消息，用于回顾上下文。"""
        import os as _os
        if not name or name != _os.path.basename(name) or name in (".", ".."):
            return {"error": f"非法的会话名: {name!r}"}
        from akm.agent_runtime.sessions import SessionStore
        session = SessionStore().load(name)
        if session is None:
            return {"error": f"会话不存在: {name}"}
        try:
            limit_n = int(limit)
        except (TypeError, ValueError):
            limit_n = 20
        limit_n = max(1, min(limit_n, 100))
        messages = session.get("messages") or []
        return {
            "name": session.get("name", ""),
            "model": session.get("model", ""),
            "created_at": session.get("created_at", ""),
            "updated_at": session.get("updated_at", ""),
            "message_count": len(messages),
            "messages": messages[-limit_n:],
        }

    async def tavily_search_tool(
        query: str,
        max_results: int = 5,
        search_depth: str = "basic",
    ) -> str:
        """通过 Tavily 远程 MCP 端点执行联网搜索，返回序列化结果。"""
        pool = getattr(app.state, "http_client", None)
        if pool is None or not getattr(pool, "is_route_pool", False):
            return json.dumps({"error": "HTTP 连接池未就绪"}, ensure_ascii=False)
        # 复用按路由隔离的连接池，遵循出站代理设置
        client = await pool.get_client(
            provider="tavily", key_alias="mcp", model="", api_path="mcp"
        )
        try:
            return await tavily_search(
                client,
                query=query,
                max_results=max_results,
                search_depth=search_depth,
            )
        except Exception as exc:
            logger.warning("[AgentTool] tavily_search 失败: %s", exc)
            return json.dumps({"error": str(exc)}, ensure_ascii=False)

    async def search_kb_tool(
        question: str,
        top_k: int = 5,
        embedding_model: str = "",
        reranker_model: str = "",
    ) -> str:
        """通过 markdown-kb 插件 HTTP 接口检索知识库，返回精选命中片段。

        以 HTTP 方式请求本机服务的 /api/markdown-kb/query 端点，与插件内部
        逻辑解耦。为控制模型上下文体积，每个命中只回传标题、文件名、相关度
        分数与截断后的正文摘要（前 500 字符），不会把完整 chunk 全量塞回。
        Agent 请求存在有效工作区时，检索范围固定为当前工作区，避免模型通过
        参数扩大到其他工作域；未配置工作区时仍可检索公共索引。
        """
        question = str(question or "").strip()
        if not question:
            return json.dumps({"error": "question 不能为空"}, ensure_ascii=False)
        payload: dict[str, Any] = {"question": question}
        top_k = max(1, min(int(top_k or 5), 20))
        payload["top_k"] = top_k
        if str(embedding_model or "").strip():
            payload["embedding_model"] = str(embedding_model).strip()
        if str(reranker_model or "").strip():
            payload["reranker_model"] = str(reranker_model).strip()
        if _request_workspace_root.get() is not None:
            try:
                workspace_root = _workspace_root()
            except ValueError as exc:
                return json.dumps({"error": str(exc)}, ensure_ascii=False)
            if workspace_root is not None:
                payload["workspace_root"] = str(workspace_root)
            else:
                # AgentLoop 即使未传 workspace_root 也会设置空的请求上下文。
                # 此时沿用公共索引检索，不能因此禁用原本无需工作区的知识库。
                payload["ignore_workspace"] = True
        else:
            payload["ignore_workspace"] = True
        url = f"{_akm_api_base_url()}/api/markdown-kb/query"
        try:
            async with httpx.AsyncClient(timeout=120.0) as client:
                resp = await client.post(url, json=payload)
                resp.raise_for_status()
                data = resp.json()
        except Exception as exc:
            logger.warning("[AgentTool] akm_search_kb 失败: %s", exc)
            return json.dumps({"error": str(exc)}, ensure_ascii=False)
        if not data.get("ok"):
            err = data.get("error") or data.get("message") or "知识库查询失败"
            return json.dumps({"error": err}, ensure_ascii=False)
        hits = data.get("hits") or []
        results = []
        for hit in hits:
            text = str(hit.get("chunk_text") or "").strip()
            results.append(
                {
                    "title": hit.get("title") or hit.get("file_name") or "",
                    "file_name": hit.get("file_name") or "",
                    "score": hit.get("score") or hit.get("vector_score") or 0,
                    "content": text[:500],
                }
            )
        return json.dumps({"results": results}, ensure_ascii=False)

    async def generate_image_tool(
        prompt: str,
        model: str = "",
        size: str = "",
        quality: str = "",
        n: int = 1,
    ) -> str:
        """调用上游图片生成接口生成图片，返回图片资源列表。

        复用连接池并走 forward_request，与 /v1/images/generations 端点共用
        同一套 key 选择与故障切换逻辑。模型未指定时取 image_supported_models
        首项，保证能匹配到可用 Key。为避免把体积巨大的 base64 数据塞进模型
        上下文，只回传 URL；生成成功后还会把图片下载保存到 agent_upload_dir
        （默认 ~/.akm/cache），并在结果中附带 local_path 与 http_url
        （http://127.0.0.1:{port}/agent-uploads/{filename}）指向该资源；
        保存失败时附带 save_error 说明原因。
        """
        pool = getattr(app.state, "http_client", None)
        if pool is None or not getattr(pool, "is_route_pool", False):
            return json.dumps({"error": "HTTP 连接池未就绪"}, ensure_ascii=False)
        cfg = load_config()
        if not str(model or "").strip():
            model = _image_default_model()
        body: dict[str, Any] = {"model": model, "prompt": prompt}
        if size:
            body["size"] = size
        if quality:
            body["quality"] = quality
        n = max(1, min(int(n or 1), _AGENT_IMAGE_MAX_COUNT))
        if n > 1:
            body["n"] = n
        client = await pool.get_client(
            provider="", key_alias="", model=model, api_path="images/generations"
        )
        try:
            # 函数内导入避免 akm.proxy 与 agent_runtime 之间循环依赖
            from akm.proxy import forward_request

            result = await forward_request(
                body,
                client,
                api_path="images/generations",
                request_timeout=_image_request_timeout(),
            )
        except Exception as exc:
            logger.warning("[AgentTool] akm_generate_image 失败: %s", exc)
            return json.dumps({"error": str(exc)}, ensure_ascii=False)
        if result.get("error") or int(result.get("status_code", 0) or 0) >= 400:
            err = result.get("error") or f"HTTP {result.get('status_code')}"
            return json.dumps({"error": err}, ensure_ascii=False)
        try:
            data = json.loads(result.get("body") or "{}")
        except (TypeError, ValueError):
            return json.dumps({"error": "上游响应不是合法 JSON"}, ensure_ascii=False)
        images = []
        for index, item in enumerate(data.get("data") or []):
            entry: dict[str, Any] = {"index": index}
            url = item.get("url")
            if url:
                entry["url"] = url
            else:
                b64 = item.get("b64_json")
                entry["b64_json_hint"] = (
                    f"base64 数据，长度 {len(b64)}" if b64 else "无可用数据"
                )
                if b64:
                    # 仅在内部交给落盘函数，_save_generated_image 会立即移除，
                    # 不能把原始 base64 回传到模型上下文。
                    entry["b64_json"] = b64
            await _save_generated_image(entry, client)
            images.append(entry)
        return json.dumps({"images": images}, ensure_ascii=False)

    async def edit_image_tool(
        prompt: str,
        image_path: str = "",
        image_base64: str = "",
        model: str = "",
        mask_path: str = "",
        mask_base64: str = "",
        size: str = "",
        quality: str = "",
        output_format: str = "",
        n: int = 1,
    ) -> str:
        """读取本地图片（或 base64 数据）并按提示词编辑，返回编辑后的图片资源列表。

        与 akm_generate_image 共用图片转发链路（forward_request + images/edits）。
        图片有两种来源：image_path 读取服务器本地文件，或 image_base64 直接传入
        base64 数据（兼容 data URL 前缀，适用于本地无文件的云端场景，模型可直接
        使用对话中 image_url 的 data URL）。handler 组装与 /v1/images/edits 一致的
        multipart 请求体。编辑结果同样会下载保存到 agent_upload_dir，并附带
        local_path 与 http_url 指向资源，保存失败时附带 save_error 说明原因。
        """
        pool = getattr(app.state, "http_client", None)
        if pool is None or not getattr(pool, "is_route_pool", False):
            return json.dumps({"error": "HTTP 连接池未就绪"}, ensure_ascii=False)
        if not str(model or "").strip():
            model = _image_default_model()
        try:
            image_file = _resolve_image(image_path, image_base64)
        except (OSError, ValueError) as exc:
            return json.dumps({"error": str(exc)}, ensure_ascii=False)
        fields: dict[str, str] = {"prompt": prompt, "model": model}
        if size:
            fields["size"] = size
        if quality:
            fields["quality"] = quality
        if output_format:
            fields["output_format"] = output_format
        n = max(1, min(int(n or 1), _AGENT_IMAGE_MAX_COUNT))
        if n > 1:
            fields["n"] = str(n)
        files: dict[str, tuple[str, bytes, str]] = {"image": image_file}
        if str(mask_path or "").strip() or str(mask_base64 or "").strip():
            try:
                files["mask"] = _resolve_image(mask_path, mask_base64)
            except (OSError, ValueError) as exc:
                return json.dumps({"error": str(exc)}, ensure_ascii=False)
        body: dict[str, Any] = {
            "__akm_multipart__": True,
            "__akm_form_fields__": fields,
            "__akm_form_files__": files,
            "model": model,
        }
        client = await pool.get_client(
            provider="", key_alias="", model=model, api_path="images/edits"
        )
        try:
            # 函数内导入避免 akm.proxy 与 agent_runtime 之间循环依赖
            from akm.proxy import forward_request

            result = await forward_request(
                body,
                client,
                api_path="images/edits",
                request_timeout=_image_request_timeout(),
            )
        except Exception as exc:
            logger.warning("[AgentTool] akm_edit_image 失败: %s", exc)
            return json.dumps({"error": str(exc)}, ensure_ascii=False)
        if result.get("error") or int(result.get("status_code", 0) or 0) >= 400:
            err = result.get("error") or f"HTTP {result.get('status_code')}"
            return json.dumps({"error": err}, ensure_ascii=False)
        try:
            data = json.loads(result.get("body") or "{}")
        except (TypeError, ValueError):
            return json.dumps({"error": "上游响应不是合法 JSON"}, ensure_ascii=False)
        images = []
        for index, item in enumerate(data.get("data") or []):
            entry: dict[str, Any] = {"index": index}
            url = item.get("url")
            if url:
                entry["url"] = url
            else:
                b64 = item.get("b64_json")
                entry["b64_json_hint"] = (
                    f"base64 数据，长度 {len(b64)}" if b64 else "无可用数据"
                )
                if b64:
                    # 同生成工具：只用于本地保存，完成后不保留在工具响应。
                    entry["b64_json"] = b64
            await _save_generated_image(entry, client)
            images.append(entry)
        return json.dumps({"images": images}, ensure_ascii=False)

    async def send_email_tool(to: str, subject: str, body: str, from_: str = "") -> dict[str, Any]:
        """发送邮件（SMTP）。返回发信结果与 Message-ID。"""
        await _check_tool_enabled("agent_email_enabled", "akm_send_email")
        # 校验收件人邮箱格式（parseaddr 只做基本解析，此处显式要求含 @）
        _to_name, _to_addr = parseaddr(to)
        if not _to_addr or "@" not in _to_addr:
            return {"ok": False, "error": "收件人邮箱格式不合法"}
        if len(body.encode("utf-8")) > _WORKSPACE_WRITE_MAX_BYTES:
            return {
                "ok": False,
                "error": f"邮件正文过大，单次上限 {_WORKSPACE_WRITE_MAX_BYTES} 字节",
            }
        cfg = load_config()
        host = str(cfg.get("agent_email_smtp_host") or "").strip()
        user = str(cfg.get("agent_email_smtp_user") or "").strip()
        if not host or not user:
            return {
                "ok": False,
                "error": "SMTP 未配置：请在 config.json 中设置 agent_email_smtp_host 与 agent_email_smtp_user 后重启服务",
            }
        try:
            port = int(cfg.get("agent_email_smtp_port") or 465)
        except (TypeError, ValueError):
            port = 465
        password = str(cfg.get("agent_email_smtp_password") or "")
        use_ssl = bool(cfg.get("agent_email_smtp_ssl", True))
        # 发件人：优先工具传入 from_，其次 SMTP 账号；同样做格式校验
        _from_name, _from_addr = parseaddr(str(from_ or "").strip())
        sender = _from_addr if (_from_addr and "@" in _from_addr) else user
        if not sender or "@" not in sender:
            return {"ok": False, "error": "发件人邮箱格式不合法"}

        def _send_sync() -> str:
            """在独立线程中执行同步 SMTP 发送，避免阻塞事件循环。"""
            msg = EmailMessage()
            msg["From"] = formataddr((_from_name, sender))
            msg["To"] = formataddr((_to_name, _to_addr))
            msg["Subject"] = subject
            msg["Message-ID"] = make_msgid()
            msg.set_content(body)
            try:
                if use_ssl:
                    server = smtplib.SMTP_SSL(host, port, timeout=30)
                else:
                    server = smtplib.SMTP(host, port, timeout=30)
                    server.starttls()
            except Exception as exc:
                raise RuntimeError(f"连接 SMTP 服务器失败: {exc}") from exc
            try:
                server.login(user, password)
                server.send_message(msg)
            finally:
                try:
                    server.quit()
                except Exception:
                    pass
            return str(msg["Message-ID"])

        try:
            message_id = await asyncio.to_thread(_send_sync)
        except Exception as exc:
            logger.warning("[AgentTool] akm_send_email 发送失败: %s", exc)
            return {"ok": False, "error": f"邮件发送失败: {exc}"}
        return {"ok": True, "message_id": message_id, "from": sender, "to": _to_addr, "subject": subject}

    empty_object = {"type": "object", "properties": {}}
    tools: list[ToolDef] = [
        ToolDef("akm_get_status", "读取 AKM 服务健康、审计队列和插件运行状态", empty_object, get_status),
        ToolDef("akm_list_keys", "列出 AKM 中已配置 Key 的非敏感状态与模型信息，不返回密钥", empty_object, get_keys),
        ToolDef("akm_get_keys_summary", "返回 AKM 当前已配置 Key 的总数，以及每个 Key 的供应商与模型清单，不返回密钥", empty_object, get_keys_summary),
        ToolDef("akm_get_time", "获取服务器当前时间，返回本地 ISO 时间、UTC 时间、UNIX 时间戳与时区", empty_object, get_time),
        ToolDef(
            "akm_get_config",
            "读取 AKM 运行配置。密钥类字段（agent_api_token、tavily_api_key）不做明文透出，仅标记是否已配置；其余配置项原样返回",
            empty_object,
            get_config,
        ),
        ToolDef(
            "akm_list_plugins",
            "列出 AKM 已加载插件的非敏感摘要：名称、版本、分类、描述、是否内置、是否启用与来源",
            empty_object,
            list_plugins,
        ),
        ToolDef(
            "akm_list_sessions",
            "列出历史 Agent 会话的元信息（会话名、创建/更新时间、消息数、模型），不含消息正文，按更新时间倒序",
            empty_object,
            list_sessions,
        ),
        ToolDef(
            "akm_load_session",
            "读取历史 Agent 会话的最近若干条消息，用于回顾之前会话的上下文",
            {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "会话名（来自 akm_list_sessions 的 name 字段）"},
                    "limit": {"type": "integer", "description": "返回最近的消息条数，1 到 100，默认 20"},
                },
                "required": ["name"],
            },
            load_session,
        ),
        ToolDef(
            "akm_list_logs",
            "查询近期 AKM 审计日志摘要，不返回请求体、响应体或请求头",
            {
                "type": "object",
                "properties": {
                    "limit": {"type": "integer", "description": "返回条数，1 到 50，默认 20"},
                    "status": {"type": "string", "enum": ["all", "success", "failed"], "description": "状态筛选，默认 all"},
                    "days": {"type": "integer", "description": "最近自然日范围，0 表示不限制，默认 1"},
                    "key_alias": {"type": "string", "description": "按 Key 别名筛选，可选"},
                },
            },
            get_logs,
        ),
        ToolDef(
            "akm_get_usage_stats",
            "查询 AKM 近期 Token 用量统计。默认同时返回最近 1/7/30 天窗口的请求数、"
            "prompt/completion/total/cached tokens，以及按 model/provider/key 的汇总。"
            "开启 cost_stats_enabled 时额外返回费用估算（total_cost）与模型单价表"
            "（input/input_cache/output，单位 USD per 1M tokens）；费用为本地估算，不能替代供应商账单",
            {
                "type": "object",
                "properties": {
                    "days": {
                        "type": "integer",
                        "description": "查询窗口：1 / 7 / 30 只返回该窗口；0 或省略时返回 1、7、30 三个窗口",
                        "enum": [0, 1, 7, 30],
                    },
                },
            },
            get_usage_stats,
        ),
        ToolDef(
            "tavily_search",
            "通过 Tavily 实时联网搜索互联网信息，返回包含标题、链接和摘要的搜索结果。需要先在 config.json 中配置 tavily_api_key",
            {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "搜索关键词"},
                    "max_results": {"type": "integer", "description": "返回结果数量，1 到 20，默认 5"},
                    "search_depth": {"type": "string", "enum": ["basic", "advanced"], "description": "搜索深度，默认 basic"},
                },
                "required": ["query"],
            },
            tavily_search_tool,
        ),
        ToolDef(
            "akm_search_kb",
            "通过 markdown-kb 知识库检索与问题最相关的文档片段，返回命中内容的标题、文件名、相关度分数与正文摘要。需要 markdown-kb 插件已启用且已学习文档",
            {
                "type": "object",
                "properties": {
                    "question": {"type": "string", "description": "检索问题"},
                    "top_k": {"type": "integer", "description": "返回命中条数，1 到 20，默认 5"},
                    "embedding_model": {"type": "string", "description": "向量模型，默认取插件配置"},
                    "reranker_model": {"type": "string", "description": "重排模型，默认取插件配置"},
                },
                "required": ["question"],
            },
            search_kb_tool,
        ),
        ToolDef(
            "akm_generate_image",
            "调用 AKM 配置的图片生成模型生成图片，返回图片资源列表。每项含 url，并附带保存到本地的 local_path 与可访问的 http_url（/agent-uploads/...），保存失败时含 save_error。需要配置 image_supported_models 对应的可用 API Key",
            {
                "type": "object",
                "properties": {
                    "prompt": {"type": "string", "description": "图片描述提示词"},
                    "model": {"type": "string", "description": "图片生成模型，默认取 image_supported_models 首项"},
                    "size": {"type": "string", "description": "图片尺寸，如 1024x1024，可选"},
                    "quality": {"type": "string", "description": "生成质量，如 standard 或 hd，可选"},
                    "n": {"type": "integer", "description": "生成张数，默认 1"},
                },
                "required": ["prompt"],
            },
            generate_image_tool,
        ),
        ToolDef(
            "akm_edit_image",
            "编辑图片（如重绘局部、扩展内容），返回编辑后的图片资源列表。每项含 url，并附带保存到本地的 local_path 与可访问的 http_url（/agent-uploads/...），保存失败时含 save_error。图片来源二选一：image_path 传服务器本地文件绝对路径；或 image_base64 传图片的 base64 数据（可直接使用对话中图片的 data:image/...;base64, 前缀数据，适合本地无文件的场景）。需要配置了对应模型的可用 API Key",
            {
                "type": "object",
                "properties": {
                    "image_path": {"type": "string", "description": "本地图片文件的绝对路径，与 image_base64 二选一"},
                    "image_base64": {"type": "string", "description": "图片 base64 数据（可带 data:image/...;base64, 前缀），与 image_path 二选一，优先级更高"},
                    "prompt": {"type": "string", "description": "编辑指令，描述期望的修改效果"},
                    "model": {"type": "string", "description": "图片编辑模型，默认取 image_supported_models 首项"},
                    "mask_path": {"type": "string", "description": "本地蒙版图片路径，用于限定重绘区域，可选"},
                    "mask_base64": {"type": "string", "description": "蒙版图片的 base64 数据（可带 data:... 前缀），与 mask_path 二选一，优先级更高"},
                    "size": {"type": "string", "description": "输出图片尺寸，如 1024x1024，可选"},
                    "quality": {"type": "string", "description": "生成质量，如 standard 或 hd，可选"},
                    "output_format": {"type": "string", "description": "输出格式，如 png 或 jpeg，可选"},
                    "n": {"type": "integer", "description": "生成张数，默认 1"},
                },
                "required": ["prompt"],
            },
            edit_image_tool,
        ),
    ]
    if load_config().get("agent_email_enabled"):
        tools.append(
            ToolDef(
                "akm_send_email",
                "发送邮件（SMTP）。需要管理员在 config.json 中配置 agent_email_smtp_host/user/password 且 agent_email_enabled=true。用于向指定邮箱发送纯文本通知",
                {
                    "type": "object",
                    "properties": {
                        "to": {"type": "string", "description": "收件人邮箱地址"},
                        "subject": {"type": "string", "description": "邮件主题"},
                        "body": {"type": "string", "description": "邮件正文（纯文本）"},
                        "from_": {"type": "string", "description": "可选发件人地址，留空使用 SMTP 账号"},
                    },
                    "required": ["to", "subject", "body"],
                },
                send_email_tool,
            )
        )
    return tools


def build_workspace_tools() -> list[ToolDef]:
    """创建工作区文件工具（读 + 写 + shell）。

    安全设计：
    - 读工具（read/list/glob/grep/info）始终注册，但目标必须位于
      agent_workspace_root 工作区内；未配置工作区时执行返回明确错误。
    - 写工具（write/edit/make_dir/delete/xlsx）默认禁用，仅当 config.json
      设置 ``agent_write_tools_enabled=true`` 时才注册，模型不可见即不可调。
    - shell 工具默认禁用，仅当 ``agent_run_shell_enabled=true`` 时才注册；
      命令由模型直接传入，系统 shell 解释执行（cwd 固定为工作区根目录）。
    """
    tools: list[ToolDef] = [
        ToolDef(
            "akm_read_file",
            "读取工作区内文本文件的内容（带行级分页与长度限制）。仅能访问 agent_workspace_root 配置的工作区目录",
            {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "工作区内的文件路径（绝对路径或相对工作区根目录的路径）"},
                    "offset": {"type": "integer", "description": "起始行号（从 0 开始），默认 0"},
                    "limit": {"type": "integer", "description": "返回的最大行数，-1 表示读到结尾，默认 -1"},
                },
                "required": ["path"],
            },
            _read_file_tool,
        ),
        ToolDef(
            "akm_list_dir",
            "列出工作区内目录下的条目（名称、类型、大小），用于感知工作区结构",
            {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "工作区内的目录路径，留空表示工作区根目录"},
                },
            },
            _list_dir_tool,
        ),
        ToolDef(
            "akm_glob",
            "在工作区内按 glob 模式匹配文件或目录路径（相对工作区根目录返回），如 **/*.py",
            {
                "type": "object",
                "properties": {
                    "pattern": {"type": "string", "description": "glob 匹配模式，如 **/*.py"},
                },
                "required": ["pattern"],
            },
            _glob_tool,
        ),
        ToolDef(
            "akm_grep",
            "在工作区内按正则搜索文件内容，返回命中的文件、行号与行内容",
            {
                "type": "object",
                "properties": {
                    "pattern": {"type": "string", "description": "要搜索的正则表达式"},
                    "path": {"type": "string", "description": "限定搜索的目录或文件（工作区内），留空递归搜索整个工作区"},
                    "case_sensitive": {"type": "boolean", "description": "是否区分大小写，默认 false"},
                },
                "required": ["pattern"],
            },
            _grep_tool,
        ),
        ToolDef(
            "akm_file_info",
            "返回工作区内文件或目录的元信息（类型、大小、修改时间）",
            {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "工作区内的文件或目录路径"},
                },
                "required": ["path"],
            },
            _file_info_tool,
        ),
    ]

    # 写工具与 shell 工具：默认禁用，仅在对应配置开关开启时注册
    if load_config().get("agent_write_tools_enabled"):
        tools.extend([
            ToolDef(
                "akm_write_file",
                "新建或覆盖工作区内的文本文件（mode=overwrite 覆盖 / append 追加）。仅 agent_write_tools_enabled=true 时可用",
                {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "工作区内的文件路径（相对工作区根目录的路径）"},
                        "content": {"type": "string", "description": "要写入的文本内容"},
                        "mode": {"type": "string", "enum": ["overwrite", "append"], "description": "overwrite 覆盖（默认）或 append 追加"},
                    },
                    "required": ["path", "content"],
                },
                _write_file_tool,
            ),
            ToolDef(
                "akm_edit_file",
                "结构化编辑工作区内的文本文件，支持两种定位方式：行号模式传 start_line（1-based，可配 end_line）把行区间替换为 new_content，推荐先读文件拿到行号后用，另可配 old_string 做锚点校验防止改错位置；内容模式将 old_string 替换为 new_string。仅 agent_write_tools_enabled=true 时可用",
                {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "工作区内的文件路径"},
                        "old_string": {"type": "string", "description": "行号模式下作为目标行区间的锚点校验内容（可选）；内容模式下为要被替换的原文片段"},
                        "new_string": {"type": "string", "description": "内容模式下的替换文本，可留空表示删除（仅内容模式使用）"},
                        "replace_all": {"type": "boolean", "description": "内容模式下是否替换所有匹配（默认只替换第一处）"},
                        "start_line": {"type": "integer", "description": "行号模式起始行号（1-based）；传此字段进入行号模式，把 [start_line, end_line] 区间整体替换为 new_content"},
                        "end_line": {"type": "integer", "description": "行号模式结束行号（含，默认等于 start_line，只替换一行）"},
                        "new_content": {"type": "string", "description": "行号模式下的新内容（可多行），替换目标行区间"},
                    },
                    "required": ["path"],
                },
                _edit_file_tool,
            ),
            ToolDef(
                "akm_make_dir",
                "在工作区内创建目录（含所需父目录）。仅 agent_write_tools_enabled=true 时可用",
                {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "工作区内的目录路径"},
                    },
                    "required": ["path"],
                },
                _make_dir_tool,
            ),
            ToolDef(
                "akm_delete_file",
                "删除工作区内的文件或目录。recursive=false（默认）只删除单个文件；recursive=true 可删除目录并递归清除其中所有内容。禁止删除工作区根目录。仅 agent_write_tools_enabled=true 时可用",
                {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "工作区内要删除的文件或目录路径"},
                        "recursive": {"type": "boolean", "description": "是否递归删除目录（含其中所有文件），默认 false"},
                    },
                    "required": ["path"],
                },
                _delete_tool,
            ),
            ToolDef(
                "akm_xlsx",
                "创建或修改工作区内的 .xlsx 电子表格文件。action=create 新建（data 为二维数组或 {sheet名: 二维数组}，目标已存在时需 overwrite=true）；action=edit 修改已有文件（updates 为 [{sheet, cell, value}] 单元格写入列表）。两种 action 共用可选自定义参数：styles 设置单元格字体/背景/对齐/数字格式，column_widths / row_heights 设列宽行高，merge_cells 合并单元格，freeze_panes 冻结窗格，charts 添加柱状/折线/饼图等图表；value 以 = 开头按公式写入。仅 agent_write_tools_enabled=true 时可用",
                {
                    "type": "object",
                    "properties": {
                        "action": {"type": "string", "enum": ["create", "edit"], "description": "create 新建 / edit 修改"},
                        "path": {"type": "string", "description": "工作区内的 .xlsx 文件路径"},
                        "data": {"type": ["array", "object"], "description": "create 用：二维数组（[[...],[...]]）或 {sheet名: 二维数组} 映射"},
                        "sheet": {"type": "string", "description": "create 时纯数组数据写入的工作表名，默认 Sheet1"},
                        "overwrite": {"type": "boolean", "description": "create 时目标已存在是否覆盖，默认 false"},
                        "updates": {"type": "array", "items": {"type": "object"}, "description": "edit 用：[{sheet, cell, value}] 单元格写入列表，cell 如 A1；value 以 = 开头写公式"},
                        "styles": {
                            "type": "array",
                            "items": {"type": "object"},
                            "description": "单元格样式：[{sheet?, cell, bold?, italic?, size?, color?, fill?, align?, number_format?}]，color/fill 为十六进制色值（如 FF0000）",
                        },
                        "column_widths": {
                            "type": "object",
                            "description": "列宽：{sheet?: {列名: 宽度}}，如 {\"Sheet1\": {\"A\": 20}}，sheet 键可省略默认 Sheet1",
                        },
                        "row_heights": {
                            "type": "object",
                            "description": "行高：{sheet?: {行号: 高度}}",
                        },
                        "merge_cells": {
                            "type": "object",
                            "description": "合并单元格：{sheet?: [区间...]}，如 {\"Sheet1\": [\"A1:C1\"]}",
                        },
                        "freeze_panes": {
                            "type": "object",
                            "description": "冻结窗格：{sheet?: \"A2\"}",
                        },
                        "charts": {
                            "type": "array",
                            "items": {"type": "object"},
                            "description": "图表列表：[{sheet?, type, title?, data_range, categories_range?, x_title?, y_title?, anchor?, legend?}]，type 为 bar/line/pie/scatter/area/doughnut，data_range 如 B2:B6",
                        },
                    },
                    "required": ["action", "path"],
                },
                _xlsx_tool,
            ),
        ])
    if load_config().get("agent_run_shell_enabled"):
        tools.append(ToolDef(
            "akm_run_shell",
            "在工作区用 shell 执行命令字符串并返回 stdout+stderr（支持管道、通配符、重定向；cwd 固定为工作区根目录）。仅 agent_run_shell_enabled=true 时可用；属主机级进程执行能力，需配合插件策略使用",
            {
                "type": "object",
                "properties": {
                    "command": {"type": "string", "description": "要执行的 shell 命令字符串"},
                    "timeout": {"type": "integer", "description": "超时秒数，1-300，默认 60"},
                },
                "required": ["command"],
            },
            _run_shell_tool,
        ))
    if load_config().get("agent_git_enabled"):
        tools.append(ToolDef(
            "akm_run_git",
            "在工作区执行结构化 git 操作并返回输出与退出码。仅支持 status、diff、log、show、add、restore、reset、commit、branch；不接受自由命令字符串。仅 agent_git_enabled=true 时可用",
            {
                "type": "object",
                "properties": {
                    "operation": {"type": "string", "enum": ["status", "diff", "log", "show", "add", "restore", "reset", "commit", "branch"], "description": "要执行的 git 操作"},
                    "paths": {"type": "array", "items": {"type": "string"}, "description": "add、restore、reset 必填；diff 可选的工作区相对路径"},
                    "message": {"type": "string", "description": "commit 操作必填的提交说明"},
                    "revision": {"type": "string", "description": "show 操作的 revision，默认 HEAD"},
                    "staged": {"type": "boolean", "description": "diff 是否查看暂存区，默认 false"},
                    "limit": {"type": "integer", "description": "log 返回条数，1-100，默认 20"},
                    "timeout": {"type": "integer", "description": "超时秒数，1-300，默认 60"},
                },
                "required": ["operation"],
            },
            _run_git_tool,
        ))
    return tools
